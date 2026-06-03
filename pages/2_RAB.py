import streamlit as st
from utils.supabase_client import get_supabase
from utils.helpers import format_rupiah
from datetime import datetime
from collections import defaultdict
from io import BytesIO

# ==================== IMPORT BARU ====================
from utils.ahsp_helper import get_ahsp_for_selection

# Library untuk Export
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import cm

supabase = get_supabase()
project_id = st.session_state.get("current_project_id")

st.header("📋 RAB - Rencana Anggaran Biaya")

if not project_id:
    st.warning("Pilih proyek terlebih dahulu di sidebar")
    st.stop()

# ==================== AMBIL DATA ====================
all_rab_items = supabase.table("rab_items")\
    .select("*")\
    .eq("project_id", project_id)\
    .order("level").order("sort_order").execute().data

# ==================== TRUE LIVE SEARCH ====================
st.markdown("### 🔍 Live Search (Update Langsung Saat Mengetik)")

search_term = st.text_input(
    "Ketik untuk mencari...",
    placeholder="Contoh: hebel, rangka, pasangan, A.1...",
    key="rab_live_search"
)

if search_term and search_term.strip() != "":
    search_lower = search_term.lower().strip()
    
    matched_ids = set()
    for item in all_rab_items:
        if search_lower in item.get('code', '').lower() or search_lower in item.get('description', '').lower():
            matched_ids.add(item['id'])
            current = item
            while current.get('parent_id'):
                matched_ids.add(current['parent_id'])
                current = next((x for x in all_rab_items if x['id'] == current['parent_id']), None)
                if current is None:
                    break

    filtered_items = [item for item in all_rab_items if item['id'] in matched_ids]
    match_count = len([i for i in filtered_items if search_lower in i.get('code','').lower() or search_lower in i.get('description','').lower()])
    st.success(f"✅ Ditemukan **{match_count} item** yang cocok dengan **'{search_term}'**")
else:
    filtered_items = all_rab_items

st.divider()

# ==================== TAMBAH ITEM BARU (LAMA) ====================
with st.expander("➕ Tambah Item BARU", expanded=False):
    col1, col2 = st.columns([1, 2])
    with col1:
        level = st.selectbox("Level", [0, 1, 2, 3], index=0)
        parent_options = ["Tidak ada (Main Item)"] + [f"{item['code']} - {item['description'][:40]}" for item in all_rab_items if item.get('level') == level-1]
        parent_choice = st.selectbox("Parent Item", parent_options)
    with col2:
        code = st.text_input("Kode Item", value="A.1")
        description = st.text_input("Uraian Pekerjaan", value="Pekerjaan ...")
    
    col_a, col_b, col_c = st.columns(3)
    with col_a:
        volume = st.number_input("Volume", value=1.0, step=0.01)
        unit = st.text_input("Satuan", value="m³")
    with col_b:
        unit_price = st.number_input("Harga Satuan (Rp)", value=100000, step=1000)
    with col_c:
        sort_order = st.number_input("Urutan", value=1, step=1)
    
    if st.button("💾 Simpan Item BARU", type="primary"):
        parent_id = None
        if parent_choice != "Tidak ada (Main Item)":
            parent_code = parent_choice.split(" - ")[0]
            parent = next((item for item in all_rab_items if item['code'] == parent_code), None)
            if parent:
                parent_id = parent['id']
        
        new_item = {
            "project_id": project_id, "code": code, "description": description,
            "volume": volume, "unit": unit, "unit_price": unit_price,
            "level": level, "parent_id": parent_id, "sort_order": sort_order
        }
        supabase.table("rab_items").insert(new_item).execute()
        st.success("✅ Item berhasil ditambahkan!")
        st.rerun()

st.divider()

# ==================== TAMBAH ITEM DARI AHSP (FITUR BARU) ====================
with st.expander("➕ Tambah Item dari Database AHSP", expanded=False):
    st.caption("Pilih item dari database AHSP. Harga akan otomatis terisi dari perhitungan terbaru.")

    ahsp_items = get_ahsp_for_selection()

    if not ahsp_items:
        st.warning("Belum ada data di Database AHSP. Silakan isi dulu di halaman Database AHSP.")
    else:
        # Format pilihan
        ahsp_options = {
            f"{item['code']} - {item['description'][:65]} ({item.get('unit', '-')})": item 
            for item in ahsp_items
        }

        selected_label = st.selectbox(
            "Pilih Item AHSP",
            options=list(ahsp_options.keys()),
            key="ahsp_select_rab"
        )

        selected_ahsp = ahsp_options[selected_label]

        # Preview
        col1, col2, col3 = st.columns(3)
        col1.metric("Kode", selected_ahsp['code'])
        col2.metric("Satuan", selected_ahsp.get('unit', '-'))
        col3.metric("Harga dari AHSP", f"Rp {selected_ahsp.get('calculated_unit_price', 0):,.0f}")

        st.divider()

        # Form input
        with st.form("form_add_from_ahsp"):
            col_a, col_b = st.columns(2)
            with col_a:
                level = st.selectbox("Level", [0, 1, 2, 3], index=0, key="ahsp_level")
                volume = st.number_input("Volume", value=1.0, step=0.01, key="ahsp_volume")
            with col_b:
                sort_order = st.number_input("Urutan", value=1, step=1, key="ahsp_sort")
                parent_options = ["Tidak ada (Main Item)"] + [
                    f"{item['code']} - {item['description'][:40]}" 
                    for item in all_rab_items if item.get('level') == level - 1
                ]
                parent_choice = st.selectbox("Parent Item", parent_options, key="ahsp_parent")

            if st.form_submit_button("💾 Simpan ke RAB dari AHSP", type="primary", use_container_width=True):
                try:
                    parent_id = None
                    if parent_choice != "Tidak ada (Main Item)":
                        parent_code = parent_choice.split(" - ")[0]
                        parent = next((item for item in all_rab_items if item['code'] == parent_code), None)
                        if parent:
                            parent_id = parent['id']

                    new_item = {
                        "project_id": project_id,
                        "code": selected_ahsp['code'],
                        "description": selected_ahsp['description'],
                        "volume": volume,
                        "unit": selected_ahsp.get('unit', ''),
                        "unit_price": selected_ahsp.get('calculated_unit_price', 0) or selected_ahsp.get('stored_unit_price', 0),
                        "level": level,
                        "parent_id": parent_id,
                        "sort_order": sort_order
                    }

                    supabase.table("rab_items").insert(new_item).execute()
                    st.success(f"✅ Item dari AHSP berhasil ditambahkan: {selected_ahsp['code']}")
                    st.rerun()

                except Exception as e:
                    st.error(f"Gagal menyimpan: {str(e)}")

st.divider()

# ==================== STRUKTUR RAB ====================
st.subheader("Struktur RAB")

def display_rab_tree(items, parent_id=None, level=0):
    children = [item for item in items if item.get("parent_id") == parent_id]
    for item in sorted(children, key=lambda x: x.get('sort_order', 0)):
        indent = "　" * level * 3
        total = (item.get("volume") or 0) * (item.get("unit_price") or 0)
        
        is_match = False
        if search_term:
            search_lower = search_term.lower().strip()
            if search_lower in item.get('code', '').lower() or search_lower in item.get('description', '').lower():
                is_match = True
        
        label = f"{indent}{item.get('code','')} - {item.get('description','')[:65]}"
        if is_match:
            label = f"✅ {label}"

        with st.expander(label, expanded=bool(search_term)):
            col1, col2, col3 = st.columns([3,2,2])
            col1.write(f"**Volume:** {item.get('volume','0')} {item.get('unit','')}")
            col2.write(f"**Harga Satuan:** {format_rupiah(item.get('unit_price',0))}")
            col3.write(f"**Total:** {format_rupiah(total)}")

            col_edit, col_delete = st.columns(2)
            with col_edit:
                if st.button("✏️ Edit", key=f"edit_{item['id']}"):
                    st.session_state.edit_item = item
                    st.rerun()
            with col_delete:
                if st.button("🗑️ Hapus", key=f"del_{item['id']}"):
                    st.session_state.delete_item = item
                    st.rerun()

            display_rab_tree(items, item["id"], level + 1)

if filtered_items:
    display_rab_tree(filtered_items)
else:
    if search_term:
        st.warning(f"Tidak ada item yang cocok dengan **'{search_term}'**")
    else:
        st.info("Belum ada data RAB.")

# ==================== EDIT FORM ====================
if "edit_item" in st.session_state:
    item = st.session_state.edit_item
    st.subheader(f"✏️ Edit Item: {item['code']}")
    
    with st.form("edit_rab_form"):
        col1, col2 = st.columns(2)
        with col1:
            new_code = st.text_input("Kode", value=item.get('code', ''))
            new_desc = st.text_input("Uraian Pekerjaan", value=item.get('description', ''))
            new_level = st.selectbox("Level", [0,1,2,3], index=item.get('level', 0))
        with col2:
            new_volume = st.number_input("Volume", value=float(item.get('volume', 0)), step=0.01)
            new_unit = st.text_input("Satuan", value=item.get('unit', ''))
            new_price = st.number_input("Harga Satuan (Rp)", value=float(item.get('unit_price', 0)), step=1000)

        col_btn1, col_btn2 = st.columns(2)
        with col_btn1:
            if st.form_submit_button("💾 Simpan Perubahan", type="primary", use_container_width=True):
                try:
                    supabase.table("rab_items").update({
                        "code": new_code,
                        "description": new_desc,
                        "level": new_level,
                        "volume": new_volume,
                        "unit": new_unit,
                        "unit_price": new_price,
                        "updated_at": datetime.now().isoformat()
                    }).eq("id", item['id']).execute()
                    
                    st.success("✅ Item berhasil diperbarui!")
                    del st.session_state.edit_item
                    st.rerun()
                except Exception as e:
                    st.error(f"Error: {e}")
        
        with col_btn2:
            if st.form_submit_button("Batal", use_container_width=True):
                del st.session_state.edit_item
                st.rerun()

st.divider()

# ==================== EXPORT FUNCTIONS - MULTI LEVEL TREE (FIXED) ====================

def build_tree(items):
    """Membangun struktur tree dari flat list berdasarkan parent_id"""
    children_map = defaultdict(list)
    for item in items:
        children_map[item.get('parent_id')].append(item)
    
    # Sort setiap level berdasarkan sort_order
    for pid in children_map:
        children_map[pid] = sorted(children_map[pid], key=lambda x: (x.get('sort_order', 0), x.get('id', 0)))
    
    return children_map


def process_tree_for_export(node, children_map, current_path, row_data, grand_total):
    """
    Rekursif memproses tree untuk export.
    - Jika node punya anak → dijadikan HEADER (hijau)
    - Jika node tidak punya anak → dijadikan baris data biasa
    """
    node_id = node['id']
    children = children_map.get(node_id, [])
    has_children = len(children) > 0

    # Buat nomor hierarkis saat ini
    current_number = ".".join(map(str, current_path))

    if has_children:
        # === HEADER / MAIN ITEM (punya sub) ===
        row_data.append({
            'type': 'header',
            'number': current_number,
            'description': node.get('description', ''),
            'unit': '',
            'volume': 0,
            'price': 0,
            'total': 0
        })
        
        # Proses anak-anaknya
        sub_total = 0
        for idx, child in enumerate(children, 1):
            new_path = current_path + [idx]
            child_result = process_tree_for_export(child, children_map, new_path, [], 0)
            
            # Tambahkan baris anak
            for r in child_result['rows']:
                row_data.append(r)
            
            sub_total += child_result['subtotal']
        
        # Tambahkan subtotal setelah semua anak
        row_data.append({
            'type': 'subtotal',
            'number': '',
            'description': 'SUBTOTAL',
            'unit': '',
            'volume': 0,
            'price': 0,
            'total': sub_total
        })
        
        return {'rows': row_data, 'subtotal': sub_total}
    
    else:
        # === LEAF ITEM (tidak punya anak) ===
        vol = node.get('volume', 0) or 0
        price = node.get('unit_price', 0) or 0
        total = vol * price
        
        row_data.append({
            'type': 'item',
            'number': current_number,
            'description': node.get('description', ''),
            'unit': node.get('unit', ''),
            'volume': vol,
            'price': price,
            'total': total
        })
        
        return {'rows': row_data, 'subtotal': total}


def export_rab_excel_hierarchical(rab_items, project_name):
    """Export RAB dengan dukungan multi-level hierarchy (Main → Sub → Sub-Sub)"""
    if not rab_items:
        st.warning("Tidak ada data RAB untuk diekspor.")
        return

    children_map = build_tree(rab_items)

    # Ambil root items (yang tidak punya parent atau parent-nya tidak ada di data)
    all_ids = {item['id'] for item in rab_items}
    root_items = [item for item in rab_items if item.get('parent_id') is None or item.get('parent_id') not in all_ids]
    root_items = sorted(root_items, key=lambda x: (x.get('sort_order', 0), x.get('id', 0)))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RAB"

    # Styling
    header_fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    section_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
    section_font = Font(bold=True, color="FFFFFF", size=11)
    subtotal_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
    grand_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Header
    ws.merge_cells('A1:F1')
    ws['A1'] = f"RENCANA ANGGARAN BIAYA (RAB) - {project_name}"
    ws['A1'].font = Font(bold=True, size=14, color="0d6efd")
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:F2')
    ws['A2'] = f"Tanggal Export: {datetime.now().strftime('%d %B %Y')}"
    ws['A2'].font = Font(italic=True, size=10)

    headers = ["No", "Uraian Pekerjaan", "Satuan", "Volume", "Harga Satuan (Rp)", "Jumlah (Rp)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    row_num = 5
    grand_total = 0
    all_rows = []

    # Proses setiap root item secara rekursif
    for idx, root in enumerate(root_items, 1):
        result = process_tree_for_export(root, children_map, [idx], [], 0)
        all_rows.extend(result['rows'])
        grand_total += result['subtotal']

    # Tulis ke Excel
    for row in all_rows:
        if row['type'] == 'header':
            # Section Header (Hijau)
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
            cell = ws.cell(row=row_num, column=1, value=f"▶ {row['description']}")
            cell.font = section_font
            cell.fill = section_fill
            for c in range(1, 7):
                ws.cell(row=row_num, column=c).border = thin_border
                ws.cell(row=row_num, column=c).fill = section_fill

        elif row['type'] == 'subtotal':
            ws.cell(row=row_num, column=2, value="SUBTOTAL").font = Font(bold=True)
            ws.cell(row=row_num, column=6, value=row['total']).font = Font(bold=True)
            ws.cell(row=row_num, column=6).number_format = '#,##0'
            for c in range(1, 7):
                ws.cell(row=row_num, column=c).border = thin_border
                ws.cell(row=row_num, column=c).fill = subtotal_fill

        else:  # type == 'item'
            ws.cell(row=row_num, column=1, value=row['number']).border = thin_border
            ws.cell(row=row_num, column=2, value=f"    {row['number']}. {row['description']}").border = thin_border
            ws.cell(row=row_num, column=3, value=row['unit']).border = thin_border
            ws.cell(row=row_num, column=4, value=row['volume']).border = thin_border
            ws.cell(row=row_num, column=5, value=row['price']).border = thin_border
            ws.cell(row=row_num, column=6, value=row['total']).border = thin_border

            ws.cell(row=row_num, column=4).number_format = '#,##0.00'
            ws.cell(row=row_num, column=5).number_format = '#,##0'
            ws.cell(row=row_num, column=6).number_format = '#,##0'

        row_num += 1

    # Grand Total
    row_num += 1
    ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
    ws.cell(row=row_num, column=2, value="GRAND TOTAL").font = Font(bold=True, size=11)
    ws.cell(row=row_num, column=2).fill = grand_fill
    ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='right')
    ws.cell(row=row_num, column=6, value=grand_total).font = Font(bold=True, size=11)
    ws.cell(row=row_num, column=6).fill = grand_fill
    ws.cell(row=row_num, column=6).number_format = '#,##0'
    for c in range(2, 7):
        ws.cell(row=row_num, column=c).border = thin_border
        ws.cell(row=row_num, column=c).fill = grand_fill

    # Lebar kolom
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"RAB_{project_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    st.download_button(
        label="⬇️ Download Excel (Multi-Level Hierarchy)",
        data=output,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
    st.success("✅ Export berhasil dengan struktur multi-level (Main → Sub → Detail)!")


def export_rab_pdf_hierarchical(rab_items, project_name):
    """Export RAB ke PDF dengan dukungan multi-level hierarchy"""
    if not rab_items:
        st.warning("Tidak ada data RAB untuk diekspor.")
        return

    children_map = build_tree(rab_items)

    all_ids = {item['id'] for item in rab_items}
    root_items = [item for item in rab_items if item.get('parent_id') is None or item.get('parent_id') not in all_ids]
    root_items = sorted(root_items, key=lambda x: (x.get('sort_order', 0), x.get('id', 0)))

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    normal = ParagraphStyle('Normal', parent=styles['Normal'], fontSize=8)
    title_style = ParagraphStyle('Title', parent=styles['Normal'], fontSize=13, fontName='Helvetica-Bold', textColor=colors.HexColor('#0d6efd'))
    header_style = ParagraphStyle('Header', parent=styles['Normal'], fontSize=9, fontName='Helvetica-Bold', textColor=colors.white)
    subtotal_style = ParagraphStyle('Subtotal', parent=styles['Normal'], fontSize=8, fontName='Helvetica-Bold')

    elements = []
    elements.append(Paragraph("RENCANA ANGGARAN BIAYA (RAB)", title_style))
    elements.append(Paragraph(f"<b>Proyek:</b> {project_name}", normal))
    elements.append(Paragraph(f"<b>Tanggal:</b> {datetime.now().strftime('%d %B %Y')}", normal))
    elements.append(Spacer(1, 0.3*cm))

    table_data = [["No", "Uraian Pekerjaan", "Sat", "Vol", "Harga (Rp)", "Jumlah (Rp)"]]
    grand_total = 0

    def add_to_pdf(node, path, children_map):
        nonlocal grand_total
        node_id = node['id']
        children = children_map.get(node_id, [])
        has_children = len(children) > 0
        current_no = ".".join(map(str, path))

        if has_children:
            # Header
            table_data.append(["", Paragraph(f"▶ {node.get('description', '')}", header_style), "", "", "", ""])
            
            sub_total = 0
            for idx, child in enumerate(children, 1):
                new_path = path + [idx]
                child_total = add_to_pdf(child, new_path, children_map)
                sub_total += child_total
            
            table_data.append(["", Paragraph("<b>SUBTOTAL</b>", subtotal_style), "", "", "", f"<b>{sub_total:,.0f}</b>"])
            return sub_total
        else:
            vol = node.get('volume', 0) or 0
            price = node.get('unit_price', 0) or 0
            total = vol * price
            grand_total += total

            table_data.append([
                current_no,
                Paragraph(f"    {current_no}. {node.get('description', '')}", normal),
                node.get('unit', ''),
                f"{vol:,.2f}",
                f"{price:,.0f}",
                f"{total:,.0f}"
            ])
            return total

    for idx, root in enumerate(root_items, 1):
        add_to_pdf(root, [idx], children_map)

    table_data.append(["", "GRAND TOTAL", "", "", "", f"{grand_total:,.0f}"])

    t = Table(table_data, colWidths=[1.1*cm, 7.8*cm, 1.2*cm, 1.8*cm, 2.8*cm, 2.8*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#d4edda')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#28a745')),  # first header row
    ]))
    elements.append(t)

    doc.build(elements)
    buffer.seek(0)

    filename = f"RAB_{project_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
    st.download_button(
        label="⬇️ Download PDF (Multi-Level Hierarchy)",
        data=buffer,
        file_name=filename,
        mime="application/pdf",
        use_container_width=True
    )
    st.success("✅ Export PDF multi-level berhasil!")


# ==================== TOMBOL EXPORT ====================
project_name = st.session_state.get("selected_project_name", "Proyek")

col1, col2 = st.columns(2)
with col1:
    if st.button("📊 Export ke Excel (Format Profesional)", type="primary", use_container_width=True):
        export_rab_excel_hierarchical(all_rab_items, project_name)

with col2:
    if st.button("🖨️ Export ke PDF", type="primary", use_container_width=True):
        export_rab_pdf_hierarchical(all_rab_items, project_name)
