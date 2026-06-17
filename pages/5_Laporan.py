import streamlit as st
from utils.supabase_client import get_supabase
from datetime import datetime
from io import BytesIO

supabase = get_supabase()
project_id = st.session_state.get("current_project_id")
project_name = st.session_state.get("selected_project_name", "Proyek")

st.header("🖨️ Laporan & Export")
st.subheader(f"Proyek: {project_name}")

if not project_id:
    st.warning("Pilih proyek di sidebar")
    st.stop()

st.divider()

st.subheader("📤 Export RAB ke Excel (Stabil)")

# ==================== EXPORT RAB KE EXCEL (HIRARKIS) ====================
if st.button("📊 Export ke Excel (Format Profesional)", type="primary", use_container_width=True):
    if not project_id:
        st.warning("Pilih proyek terlebih dahulu!")
        st.stop()

    rab_items = supabase.table("rab_items")\
        .select("*")\
        .eq("project_id", project_id)\
        .order("level").order("sort_order").execute().data

    if not rab_items:
        st.warning("Belum ada data RAB untuk diekspor.")
        st.stop()

    # Grouping hierarkis
    from collections import defaultdict
    children_map = defaultdict(list)
    for item in rab_items:
        children_map[item.get('parent_id')].append(item)

    # Main Items
    main_items = [item for item in rab_items 
                  if item.get('level', 0) == 0 or 
                     (item.get('volume', 0) == 0 and "pekerjaan" in item.get('description', '').lower())]

    # Buat Excel
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RAB"

    # Styling
    header_fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    main_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
    main_font = Font(bold=True, color="FFFFFF", size=11)
    subtotal_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
    grand_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Judul
    ws.merge_cells('A1:F1')
    ws['A1'] = f"RENCANA ANGGARAN BIAYA (RAB) - {project_name}"
    ws['A1'].font = Font(bold=True, size=14, color="0d6efd")
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:F2')
    ws['A2'] = f"Tanggal Export: {datetime.now().strftime('%d %B %Y')}"
    ws['A2'].font = Font(italic=True, size=10)

    # Header kolom
    headers = ["No", "Uraian Pekerjaan", "Satuan", "Volume", "Harga Satuan (Rp)", "Jumlah (Rp)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Isi data
    row_num = 5
    item_no = 1
    grand_total = 0

    for main_item in main_items:
        main_id = main_item['id']
        main_desc = main_item.get('description', '')

        # Baris Main Item (Hijau)
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
        cell = ws.cell(row=row_num, column=1, value=f"▶ {main_desc}")
        cell.font = main_font
        cell.fill = main_fill
        for col in range(1, 7):
            ws.cell(row=row_num, column=col).border = thin_border
            ws.cell(row=row_num, column=col).fill = main_fill
        row_num += 1

        main_subtotal = 0

        # Sub Items
        for child in children_map.get(main_id, []):
            vol = child.get('volume', 0) or 0
            price = child.get('unit_price', 0) or 0
            total = vol * price
            main_subtotal += total
            grand_total += total

            ws.cell(row=row_num, column=1, value=item_no).border = thin_border
            ws.cell(row=row_num, column=2, value=f"    {item_no}. {child.get('description', '')}").border = thin_border
            ws.cell(row=row_num, column=3, value=child.get('unit', '')).border = thin_border
            ws.cell(row=row_num, column=4, value=vol).border = thin_border
            ws.cell(row=row_num, column=5, value=price).border = thin_border
            ws.cell(row=row_num, column=6, value=total).border = thin_border

            ws.cell(row=row_num, column=4).number_format = '#,##0.00'
            ws.cell(row=row_num, column=5).number_format = '#,##0'
            ws.cell(row=row_num, column=6).number_format = '#,##0'

            item_no += 1
            row_num += 1

        # Subtotal per Main Item
        ws.cell(row=row_num, column=2, value="SUBTOTAL").font = Font(bold=True)
        ws.cell(row=row_num, column=6, value=main_subtotal).font = Font(bold=True)
        ws.cell(row=row_num, column=6).fill = subtotal_fill
        ws.cell(row=row_num, column=6).number_format = '#,##0'
        for col in range(1, 7):
            ws.cell(row=row_num, column=col).border = thin_border
            ws.cell(row=row_num, column=col).fill = subtotal_fill
        row_num += 1

    # Grand Total
    row_num += 1
    ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
    ws.cell(row=row_num, column=2, value="GRAND TOTAL").font = Font(bold=True, size=11)
    ws.cell(row=row_num, column=2).fill = grand_fill
    ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='right')

    ws.cell(row=row_num, column=6, value=grand_total).font = Font(bold=True, size=11)
    ws.cell(row=row_num, column=6).fill = grand_fill
    ws.cell(row=row_num, column=6).number_format = '#,##0'
    for col in range(2, 7):
        ws.cell(row=row_num, column=col).border = thin_border
        ws.cell(row=row_num, column=col).fill = grand_fill

    # Lebar kolom
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18

    # Download
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    st.download_button(
        label="⬇️ Download File Excel (Format Hirarkis)",
        data=output,
        file_name=f"RAB_{project_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
    st.success("Export RAB berhasil!")

st.divider()

st.subheader("📤 Import RAB dari Excel")

uploaded_file = st.file_uploader("Pilih file Excel (.xlsx)", type=["xlsx"])

if uploaded_file:
    try:
        import pandas as pd
        df = pd.read_excel(uploaded_file)

        required_cols = ['Level', 'Kode', 'Uraian Pekerjaan', 'Volume', 'Satuan', 'Harga Satuan']
        if not all(col in df.columns for col in required_cols):
            st.error(f"Kolom yang dibutuhkan: {required_cols}")
            st.stop()

        success_count = 0
        failed_rows = []
        for idx, (_, row) in enumerate(df.iterrows()):
            try:
                data = {
                    "project_id": project_id,
                    "level": int(row['Level']) if pd.notna(row['Level']) else 0,
                    "code": str(row['Kode']) if pd.notna(row['Kode']) else "",
                    "description": str(row['Uraian Pekerjaan']) if pd.notna(row['Uraian Pekerjaan']) else "",
                    "volume": float(row['Volume']) if pd.notna(row['Volume']) else 0,
                    "unit": str(row['Satuan']) if pd.notna(row['Satuan']) else "",
                    "unit_price": float(row['Harga Satuan']) if pd.notna(row['Harga Satuan']) else 0,
                    "parent_id": None,
                    "sort_order": success_count + 1
                }
                supabase.table("rab_items").insert(data).execute()
                success_count += 1
            except Exception as row_err:
                # Jangan telan diam-diam: kumpulkan baris yang gagal agar user tahu.
                failed_rows.append((idx + 1, str(row_err)))

        if success_count > 0:
            st.success(f"✅ Berhasil import {success_count} item dari Excel!")
            if failed_rows:
                st.warning(
                    f"⚠️ {len(failed_rows)} baris gagal diimport. "
                    "Periksa kolom/tipe data baris berikut:"
                )
                for rownum, err in failed_rows[:10]:
                    st.caption(f"Baris {rownum}: {err}")
            st.rerun()
        else:
            st.error("❌ Tidak ada data yang berhasil diimport")
            if failed_rows:
                for rownum, err in failed_rows[:10]:
                    st.caption(f"Baris {rownum}: {err}")
    except Exception as e:
        st.error(f"❌ Error: {str(e)}")

st.divider()

# ==================== EXPORT RAB KE PDF ====================
if st.button("🖨️ Export ke PDF (Format Profesional)", use_container_width=True):
    if not project_id:
        st.warning("Pilih proyek terlebih dahulu!")
        st.stop()

    rab_items = supabase.table("rab_items")\
        .select("*")\
        .eq("project_id", project_id)\
        .order("level").order("sort_order").execute().data

    if not rab_items:
        st.warning("Belum ada data RAB")
        st.stop()

    from collections import defaultdict
    children_map = defaultdict(list)
    for item in rab_items:
        children_map[item.get('parent_id')].append(item)

    main_items = [item for item in rab_items 
                  if item.get('level', 0) == 0 or 
                     (item.get('volume', 0) == 0 and "pekerjaan" in item.get('description', '').lower())]

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm
    from io import BytesIO

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    normal = ParagraphStyle('Normal', parent=styles['Normal'], fontSize=9)
    title_style = ParagraphStyle('Title', parent=styles['Normal'], fontSize=14, fontName='Helvetica-Bold', textColor=colors.HexColor('#0d6efd'))
    main_style = ParagraphStyle('Main', parent=styles['Normal'], fontSize=9, fontName='Helvetica-Bold', textColor=colors.HexColor('#2E7D32'))

    elements = []
    elements.append(Paragraph("RENCANA ANGGARAN BIAYA (RAB)", title_style))
    elements.append(Paragraph(f"<b>Proyek:</b> {project_name}", normal))
    elements.append(Paragraph(f"<b>Tanggal:</b> {datetime.now().strftime('%d %B %Y')}", normal))
    elements.append(Spacer(1, 0.4*cm))

    table_data = [["No", "Uraian Pekerjaan", "Sat", "Vol", "Harga (Rp)", "Jumlah (Rp)"]]
    item_no = 1
    grand_total = 0

    for main_item in main_items:
        main_id = main_item['id']
        table_data.append(["", Paragraph(f"▶ {main_item.get('description', '')}", main_style), "", "", "", ""])

        for child in children_map.get(main_id, []):
            vol = child.get('volume', 0) or 0
            price = child.get('unit_price', 0) or 0
            total = vol * price
            grand_total += total

            table_data.append([
                str(item_no),
                Paragraph(f"    {item_no}. {child.get('description', '')}", normal),
                child.get('unit', ''),
                f"{vol:,.2f}",
                f"{price:,.0f}",
                f"{total:,.0f}"
            ])
            item_no += 1

    table_data.append(["", "GRAND TOTAL", "", "", "", f"{grand_total:,.0f}"])

    t = Table(table_data, colWidths=[1*cm, 7*cm, 1.5*cm, 2*cm, 3*cm, 3*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#d4edda')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(t)

    doc.build(elements)
    buffer.seek(0)

    st.download_button(
        label="⬇️ Download RAB PDF",
        data=buffer,
        file_name=f"RAB_{project_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf",
        mime="application/pdf",
        use_container_width=True
    )
st.info("💡 Halaman ini menyediakan export RAB ke Excel & PDF, serta import RAB dari Excel.")

st.caption(f"Update: {datetime.now().strftime('%d %B %Y %H:%M')}")