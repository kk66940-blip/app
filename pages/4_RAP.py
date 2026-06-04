import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from utils.supabase_client import get_supabase
from utils.helpers import format_rupiah
from datetime import datetime
from io import BytesIO

supabase = get_supabase()
project_id = st.session_state.get("current_project_id")
project_name = st.session_state.get("selected_project_name", "Proyek")

st.header("📋 RAP - Rencana Anggaran Pelaksanaan")
st.subheader(f"Proyek: {project_name}")

if not project_id:
    st.warning("Pilih proyek di sidebar")
    st.stop()

st.divider()

# ==================== CREATE RAP FROM RAB ====================
st.subheader("🔄 Buat RAP dari RAB")

col1, col2 = st.columns([1, 2])
with col1:
    percentage = st.number_input(
        "Persentase Harga dari RAB (%)",
        min_value=50,
        max_value=150,
        value=85,
        step=1
    )

with col2:
    st.write("")
    st.write("")
    if st.button("🔄 Buat/Update RAP", type="primary", use_container_width=True):
        try:
            supabase.table("rap_items").delete().eq("project_id", project_id).execute()
            
            rab_items = supabase.table("rab_items")\
                .select("*")\
                .eq("project_id", project_id)\
                .execute().data
            
            for item in rab_items:
                rap_data = {
                    "project_id": project_id,
                    "rab_item_id": item['id'],
                    "code": item.get('code', ''),
                    "description": item.get('description', ''),
                    "unit": item.get('unit', ''),
                    "volume": item.get('volume', 0),
                    "planned_price": item.get('unit_price', 0),
                    "execution_price": (item.get('unit_price', 0) * percentage / 100),
                    "upah": 0,
                    "level": item.get('level', 0),
                    "parent_id": item.get('parent_id')
                }
                supabase.table("rap_items").insert(rap_data).execute()
            
            st.success("✅ RAP berhasil dibuat!")
            st.rerun()
        except Exception as e:
            st.error(f"❌ Error: {str(e)}")

st.divider()

# ==================== EXPORT BUTTONS (VERSI FINAL - SMART GROUPING) ====================
st.subheader("📤 Export RAP")

col1, col2 = st.columns(2)

with col1:
    if st.button("📊 Export ke Excel (Format Profesional)", type="primary", use_container_width=True):
        try:
            rap_items = supabase.table("rap_items")\
                .select("*")\
                .eq("project_id", project_id)\
                .execute().data

            if not rap_items:
                st.warning("Tidak ada data RAP untuk diekspor.")
                st.stop()

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "RAP"

            # ==================== STYLING ====================
            header_fill = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            main_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
            main_font = Font(bold=True, color="FFFFFF", size=11)
            
            subtotal_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
            subtotal_font = Font(bold=True, size=10)
            
            grand_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
            grand_font = Font(bold=True, color="FFFFFF", size=12)
            
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))

            # ==================== JUDUL ====================
            ws.merge_cells('A1:G1')
            ws['A1'] = "RENCANA ANGGARAN PELAKSANAAN (RAP)"
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = Alignment(horizontal='center')

            ws.merge_cells('A2:G2')
            ws['A2'] = f"Proyek: {project_name}   |   Tanggal: {datetime.now().strftime('%d/%m/%Y')}"
            ws['A2'].font = Font(size=11, italic=True)
            ws['A2'].alignment = Alignment(horizontal='center')

            # ==================== HEADER ====================
            headers = ["No", "Uraian Pekerjaan", "Sat", "Vol", "Harga Rencana (Rp)", "Harga Pelaksanaan (Rp)", "Upah (Rp)", "Total Biaya (Rp)"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            # ==================== SMART GROUPING (BERDASARKAN DESKRIPSI) ====================
            current_row = 5
            main_counter = 0
            sub_counter = 0
            current_main_total = 0
            grand_total = 0

            # Urutkan berdasarkan id (asumsi urutan sudah benar)
            sorted_items = sorted(rap_items, key=lambda x: x.get('id', 0))

            i = 0
            while i < len(sorted_items):
                item = sorted_items[i]
                desc = item.get('description', '').strip().lower()
                volume = item.get('volume') or 0

                # Deteksi main item: volume = 0 DAN deskripsi mengandung "pekerjaan"
                is_main = (volume == 0) and ("pekerjaan" in desc)

                if is_main:
                    # ==================== MAIN ITEM (HIJAU) ====================
                    main_counter += 1
                    sub_counter = 0
                    
                    ws.cell(row=current_row, column=1, value=main_counter).border = thin_border
                    ws.cell(row=current_row, column=2, value=item.get('description', '')).border = thin_border
                    ws.cell(row=current_row, column=3, value=item.get('unit', '')).border = thin_border
                    ws.cell(row=current_row, column=4, value=item.get('volume', 0)).border = thin_border
                    ws.cell(row=current_row, column=5, value=item.get('planned_price', 0)).border = thin_border
                    ws.cell(row=current_row, column=6, value=item.get('execution_price', 0)).border = thin_border
                    ws.cell(row=current_row, column=7, value=item.get('upah', 0)).border = thin_border
                    
                    total = (item.get('volume') or 0) * (item.get('execution_price') or 0)
                    ws.cell(row=current_row, column=8, value=total).border = thin_border
                    
                    for col in range(1, 8):
                        ws.cell(row=current_row, column=col).fill = main_fill
                        ws.cell(row=current_row, column=col).font = main_font
                    
                    for c in [4, 5, 6, 7]:
                        ws.cell(row=current_row, column=c).number_format = '#,##0'
                    
                    current_row += 1
                    current_main_total = total
                    grand_total += total
                    
                    # Ambil item berikutnya sebagai child (sampai ketemu main item lagi)
                    i += 1
                    while i < len(sorted_items):
                        next_item = sorted_items[i]
                        next_desc = next_item.get('description', '').strip().lower()
                        next_vol = next_item.get('volume') or 0
                        next_is_main = (next_vol == 0) and ("pekerjaan" in next_desc)
                        
                        if next_is_main:
                            break  # ketemu main item baru, stop
                        
                        # Ini child
                        sub_counter += 1
                        
                        ws.cell(row=current_row, column=1, value=f"{main_counter}.{sub_counter}").border = thin_border
                        ws.cell(row=current_row, column=2, value="    " + next_item.get('description', '')).border = thin_border
                        ws.cell(row=current_row, column=3, value=next_item.get('unit', '')).border = thin_border
                        ws.cell(row=current_row, column=4, value=next_item.get('volume', 0)).border = thin_border
                        ws.cell(row=current_row, column=5, value=next_item.get('planned_price', 0)).border = thin_border
                        ws.cell(row=current_row, column=6, value=next_item.get('execution_price', 0)).border = thin_border
                        ws.cell(row=current_row, column=7, value=next_item.get('upah', 0)).border = thin_border
                        
                        child_total = (next_item.get('volume') or 0) * (next_item.get('execution_price') or 0)
                        ws.cell(row=current_row, column=8, value=child_total).border = thin_border
                        
                        for c in [4, 5, 6, 7]:
                            ws.cell(row=current_row, column=c).number_format = '#,##0'
                        
                        current_main_total += child_total
                        grand_total += child_total
                        current_row += 1
                        i += 1
                    
                    # SUBTOTAL
                    ws.cell(row=current_row, column=2, value="SUBTOTAL").font = subtotal_font
                    ws.cell(row=current_row, column=8, value=current_main_total).font = subtotal_font
                    ws.cell(row=current_row, column=8).number_format = '#,##0'
                    
                    for col in range(2, 8):
                        ws.cell(row=current_row, column=col).fill = subtotal_fill
                        ws.cell(row=current_row, column=col).border = thin_border
                    current_row += 1
                    
                else:
                    i += 1

            # ==================== GRAND TOTAL ====================
            current_row += 1
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=7)
            ws.cell(row=current_row, column=2, value="GRAND TOTAL").font = grand_font
            ws.cell(row=current_row, column=2).fill = grand_fill
            ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='right')
            
            ws.cell(row=current_row, column=8, value=grand_total).font = grand_font
            ws.cell(row=current_row, column=8).fill = grand_fill
            ws.cell(row=current_row, column=8).number_format = '#,##0'
            
            for col in range(2, 8):
                ws.cell(row=current_row, column=col).fill = grand_fill
                ws.cell(row=current_row, column=col).border = thin_border

            # Lebar kolom
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 55
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 22
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 20

            ws.row_dimensions[1].height = 28
            ws.row_dimensions[4].height = 22

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            filename = f"RAP_{project_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            st.download_button(
                label="⬇️ Download File Excel (Format Profesional)",
                data=output,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            st.success("✅ Export berhasil! Sekarang sub-item sudah masuk ke bawah main item dengan benar.")

        except Exception as e:
            st.error(f"❌ Error: {str(e)}")

with col2:
    if st.button("📄 Export ke PDF", type="primary", use_container_width=True):
        st.info("Fitur PDF akan ditambahkan nanti.")

st.divider()

# ==================== TABEL DI HALAMAN ====================
st.subheader("📊 Daftar RAP (Hirarkis)")

rap_items = supabase.table("rap_items")\
    .select("*")\
    .eq("project_id", project_id)\
    .execute().data

if not rap_items:
    st.info("Belum ada data RAP. Buat RAP dari RAB di atas.")
    st.stop()

def display_rap_like_rab(items):
    """Tampilan RAP bersih berdasarkan Level (semua level ditampilkan)"""
    if not items:
        return

    from collections import defaultdict
    level_map = defaultdict(list)
    for item in items:
        lvl = item.get('level', 0)
        level_map[lvl].append(item)
    
    for lvl in level_map:
        level_map[lvl] = sorted(level_map[lvl], key=lambda x: (x.get('sort_order', 0), x.get('id', 0)))

    st.subheader("📊 Struktur RAP (Hierarkis berdasarkan Level)")

    for lvl in sorted(level_map.keys()):
        items_in_level = level_map[lvl]
        for item in items_in_level:
            indent = "　" * lvl
            code = item.get('code', '')
            desc = item.get('description', '')
            
            if lvl == 0:
                prefix = "▶ "
            else:
                prefix = "└─ "
            
            title = f"{indent}{prefix}{code} - {desc}" if code else f"{indent}{prefix}{desc}"

            total_rencana = (item.get("volume") or 0) * (item.get("planned_price") or 0)
            total_pelaksanaan = (item.get("volume") or 0) * (item.get("execution_price") or 0)
            total_upah = (item.get("upah") or 0) * (item.get("volume") or 0)

            with st.expander(title, expanded=False):
                col1, col2, col3 = st.columns(3)
                col1.metric("Volume", f"{item.get('volume','0')} {item.get('unit','')}")
                col2.metric("Harga Rencana", format_rupiah(item.get('planned_price',0)))
                col3.metric("Harga Pelaksanaan", format_rupiah(item.get('execution_price',0)))

                st.caption(f"**Total Rencana:** {format_rupiah(total_rencana)} | **Total Pelaksanaan:** {format_rupiah(total_pelaksanaan)} | **Total + Upah:** {format_rupiah(total_upah)}")

                col_edit, col_delete = st.columns(2)
                with col_edit:
                    if st.button("✏️ Edit Harga", key=f"edit_{item['id']}", use_container_width=True):
                        st.session_state.edit_rap_item = item
                        st.rerun()
                with col_delete:
                    if st.button("🗑️ Hapus", key=f"del_{item['id']}", use_container_width=True):
                        st.warning("Fitur hapus akan ditambahkan nanti")

    # Form Edit
    if "edit_rap_item" in st.session_state:
        item = st.session_state.edit_rap_item
        st.subheader(f"✏️ Edit Item: {item.get('code','')} - {item.get('description','')}")

        col1, col2 = st.columns(2)
        with col1:
            new_exec = st.number_input("Harga Pelaksanaan Baru (Rp)", value=float(item.get('execution_price', 0)), step=1000.0)
        with col2:
            new_upah = st.number_input("Upah Baru (Rp)", value=float(item.get('upah', 0)), step=1000.0)

        col_save, col_cancel = st.columns(2)
        with col_save:
            if st.button("💾 Simpan Perubahan", type="primary"):
                try:
                    supabase.table("rap_items").update({
                        "execution_price": new_exec,
                        "upah": new_upah
                    }).eq("id", item['id']).execute()
                    st.success("✅ Berhasil disimpan!")
                    del st.session_state.edit_rap_item
                    st.rerun()
                except Exception as e:
                    st.error(f"Error: {str(e)}")
        with col_cancel:
            if st.button("Batal"):
                del st.session_state.edit_rap_item
                st.rerun()

display_rap_like_rab(rap_items)

st.divider()

# ==================== SUMMARY ====================
st.subheader("📈 Ringkasan RAP")

total_rencana = sum(item.get('volume', 0) * item.get('planned_price', 0) for item in rap_items)
total_pelaksanaan = sum(item.get('volume', 0) * item.get('execution_price', 0) for item in rap_items)
total_upah = sum(item.get('volume', 0) * item.get('upah', 0) for item in rap_items)
total_biaya = total_pelaksanaan
total_variance = total_rencana - total_biaya

col1, col2, col3, col4 = st.columns(4)
with col1:
    st.metric("Total Rencana", format_rupiah(total_rencana))
with col2:
    st.metric("Total Pelaksanaan", format_rupiah(total_pelaksanaan))
with col3:
    st.metric("Total Upah (Info)", format_rupiah(total_upah))
with col4:
    st.metric("Total Biaya", format_rupiah(total_biaya), 
              delta=format_rupiah(total_variance),
              delta_color="inverse" if total_variance < 0 else "normal")

st.caption(f"Update: {datetime.now().strftime('%d %B %Y %H:%M')}")
