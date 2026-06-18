import streamlit as st
from utils.supabase_client import get_supabase
from utils.helpers import format_rupiah
from datetime import datetime
from collections import defaultdict
from io import BytesIO
import uuid
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

supabase = get_supabase()
project_id = st.session_state.get("current_project_id")
project_name = st.session_state.get("selected_project_name", "Proyek")

st.header("💰 Pengeluaran Proyek")
st.subheader(f"Proyek: {project_name}")

if not project_id:
    st.warning("Pilih proyek di sidebar terlebih dahulu.")
    st.stop()

# ==================== FUNGSI EXPORT EXCEL (HIRARKI PER KATEGORI) ====================
def export_expenses_to_excel(expenses, project_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Pengeluaran"

    # Styling
    header_fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    category_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
    category_font = Font(bold=True, color="FFFFFF", size=11)
    subtotal_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
    total_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    title_font = Font(bold=True, size=14, color="0d6efd")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Kop surat (gagal-aman). Mengembalikan baris awal konten.
    base = 1
    try:
        from utils.company import get_company_settings, add_excel_letterhead
        base = add_excel_letterhead(ws, get_company_settings(), num_cols=6)
    except Exception:
        base = 1

    title_row, date_row, hdr_row = base, base + 1, base + 3

    # Judul
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=6)
    tc = ws.cell(row=title_row, column=1, value=f"LAPORAN PENGELUARAN PROYEK - {project_name}")
    tc.font = title_font
    tc.alignment = Alignment(horizontal='center')

    ws.merge_cells(start_row=date_row, start_column=1, end_row=date_row, end_column=6)
    dc = ws.cell(row=date_row, column=1, value=f"Tanggal Export: {datetime.now().strftime('%d %B %Y')}")
    dc.font = Font(italic=True, size=10)

    # Header Tabel
    headers = ["No", "Tanggal", "Kategori", "Uraian", "Dibayar Oleh", "Jumlah (Rp)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=hdr_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Grouping data berdasarkan Kategori
    grouped = defaultdict(list)
    for exp in expenses:
        grouped[exp['category']].append(exp)

    current_row = hdr_row + 1
    item_no = 1
    grand_total = 0

    for category, items in grouped.items():
        # Header Kategori
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        for col in range(1, 7):
            ws.cell(row=current_row, column=col).fill = category_fill
            ws.cell(row=current_row, column=col).border = thin_border
        ws.cell(row=current_row, column=1, value=f"▶ {category}")
        ws.cell(row=current_row, column=1).font = category_font
        current_row += 1

        category_total = 0

        # Data per kategori
        for exp in items:
            ws.cell(row=current_row, column=1, value=item_no).border = thin_border
            ws.cell(row=current_row, column=2, value=exp['expense_date']).border = thin_border
            ws.cell(row=current_row, column=3, value=exp['category']).border = thin_border
            ws.cell(row=current_row, column=4, value=exp.get('description', '')).border = thin_border
            ws.cell(row=current_row, column=5, value=exp.get('paid_by', '')).border = thin_border
            ws.cell(row=current_row, column=6, value=exp.get('amount', 0)).border = thin_border
            ws.cell(row=current_row, column=6).number_format = '#,##0'

            category_total += exp.get('amount', 0)
            grand_total += exp.get('amount', 0)
            item_no += 1
            current_row += 1

        # Subtotal per Kategori
        ws.cell(row=current_row, column=5, value="SUBTOTAL").font = Font(bold=True)
        ws.cell(row=current_row, column=6, value=category_total).font = Font(bold=True)
        ws.cell(row=current_row, column=6).number_format = '#,##0'
        for col in range(1, 7):
            ws.cell(row=current_row, column=col).fill = subtotal_fill
            ws.cell(row=current_row, column=col).border = thin_border
        current_row += 1

    # Grand Total
    current_row += 1
    ws.cell(row=current_row, column=5, value="GRAND TOTAL").font = Font(bold=True, size=11)
    ws.cell(row=current_row, column=5).fill = total_fill
    ws.cell(row=current_row, column=5).alignment = Alignment(horizontal='right')
    ws.cell(row=current_row, column=6, value=grand_total).font = Font(bold=True, size=11)
    ws.cell(row=current_row, column=6).fill = total_fill
    ws.cell(row=current_row, column=6).number_format = '#,##0'
    for col in range(5, 7):
        ws.cell(row=current_row, column=col).border = thin_border

    # Lebar Kolom
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ==================== TOMBOL EXPORT ====================
expenses = supabase.table("project_expenses") \
    .select("*") \
    .eq("project_id", project_id) \
    .order("expense_date", desc=True) \
    .execute().data

col1, col2 = st.columns([3, 1])
with col2:
    if expenses and st.button("📊 Export ke Excel", type="primary", use_container_width=True):
        excel_file = export_expenses_to_excel(expenses, project_name)
        filename = f"Pengeluaran_{project_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        st.download_button(
            label="⬇️ Download Laporan Excel",
            data=excel_file,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

st.divider()

# ==================== FORM TAMBAH PENGELUARAN ====================
with st.expander("➕ Tambah Pengeluaran Baru", expanded=False):
    with st.form("form_add_expense", clear_on_submit=True):
        col1, col2 = st.columns(2)
        with col1:
            expense_date = st.date_input("Tanggal Pengeluaran", datetime.now().date())
            category = st.selectbox("Kategori Pengeluaran *", 
                ["Material", "Upah / Tenaga Kerja", "Sewa Alat", "BBM / Transportasi", "Lain-lain"])
            other_category = st.text_input("Isi Kategori Lainnya *") if category == "Lain-lain" else ""

        with col2:
            amount = st.number_input("Jumlah (Rp) *", min_value=0, step=10000, format="%d")
            paid_by = st.text_input("Dibayar Oleh")

        description = st.text_area("Uraian / Keterangan")
        notes = st.text_input("Catatan Tambahan")

        uploaded_file = st.file_uploader("Upload Bukti Foto", type=["jpg", "png", "jpeg"])

        if st.form_submit_button("💾 Simpan", type="primary", use_container_width=True):
            if amount > 0:
                try:
                    final_category = other_category if category == "Lain-lain" else category
                    data = {
                        "project_id": project_id,
                        "expense_date": str(expense_date),
                        "category": final_category,
                        "description": description,
                        "amount": amount,
                        "paid_by": paid_by,
                        "notes": notes,
                        "created_by": st.session_state.user.get("username", "unknown")
                    }
                    if uploaded_file:
                        file_path = f"expenses/{project_id}/{uuid.uuid4()}.{uploaded_file.name.split('.')[-1]}"
                        supabase.storage.from_("opname-photos").upload(file_path, uploaded_file.getvalue())
                        data["receipt_photo_url"] = supabase.storage.from_("opname-photos").get_public_url(file_path)

                    supabase.table("project_expenses").insert(data).execute()
                    st.success("Pengeluaran berhasil disimpan!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Gagal menyimpan: {e}")
            else:
                st.error("Jumlah harus lebih dari 0.")

st.divider()

# ==================== RINGKASAN & DAFTAR ====================
st.subheader("📊 Ringkasan Pengeluaran")

if expenses:
    total = sum(e.get("amount", 0) for e in expenses)
    st.metric("Total Pengeluaran", f"Rp {total:,.0f}")

    cat_sum = defaultdict(float)
    for e in expenses:
        cat_sum[e["category"]] += e.get("amount", 0)

    cols = st.columns(len(cat_sum))
    for i, (cat, val) in enumerate(cat_sum.items()):
        with cols[i]:
            st.metric(cat, f"Rp {val:,.0f}")

st.divider()
st.subheader("📋 Daftar Pengeluaran")

if not expenses:
    st.info("Belum ada data pengeluaran.")
else:
    for exp in expenses:
        with st.expander(f"{exp['expense_date']} | {exp['category']} | Rp {exp['amount']:,.0f}"):
            st.write(f"**Uraian:** {exp.get('description', '-')}")
            st.write(f"**Dibayar Oleh:** {exp.get('paid_by', '-')}")
            if exp.get('notes'):
                st.write(f"**Catatan:** {exp.get('notes')}")
            if exp.get("receipt_photo_url"):
                st.image(exp["receipt_photo_url"], width=200)

            col1, col2 = st.columns(2)
            with col1:
                if st.button("✏️ Edit", key=f"edit_{exp['id']}"):
                    st.session_state.edit_expense = exp
                    st.rerun()
            with col2:
                with st.popover("🗑️ Hapus", use_container_width=True):
                    st.write(f"Hapus pengeluaran **{exp.get('description', '')}** "
                             f"({format_rupiah(exp.get('amount', 0))})?")
                    if st.button("Ya, hapus permanen", key=f"delconfirm_{exp['id']}", type="primary"):
                        try:
                            supabase.table("project_expenses").delete().eq("id", exp['id']).execute()
                            st.success("Pengeluaran dihapus.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Gagal hapus: {e}")

# Form Edit
if "edit_expense" in st.session_state:
    exp = st.session_state.edit_expense
    with st.form("edit_form"):
        new_amount = st.number_input("Jumlah", value=float(exp['amount']), step=10000.0)
        new_paid_by = st.text_input("Dibayar Oleh", value=exp.get('paid_by', ''))
        if st.form_submit_button("Simpan Perubahan"):
            supabase.table("project_expenses").update({
                "amount": new_amount, "paid_by": new_paid_by
            }).eq("id", exp['id']).execute()
            del st.session_state.edit_expense
            st.rerun()
