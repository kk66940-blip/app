"""
pages/15_Laporan_Proyek.py
Laporan Lengkap Proyek — pilih bagian, lalu export ke PDF atau Excel.

Bagian: Header proyek + kop perusahaan, Ringkasan keuangan, Rincian pembayaran.
Memakai helper kop (build_letterhead / add_excel_letterhead) dan perhitungan
keuangan compute_project_recap yang sudah ada (konsisten dgn Rekap).
"""

import sys
from io import BytesIO
from pathlib import Path
from datetime import datetime

import streamlit as st

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from utils.supabase_client import get_supabase
from utils.helpers import format_rupiah, compute_project_recap
from utils.company import get_company_settings, build_letterhead, add_excel_letterhead

supabase = get_supabase()

st.header("📄 Laporan Lengkap Proyek")

project_id = st.session_state.get("current_project_id")
if not project_id:
    st.warning("Pilih proyek di sidebar terlebih dahulu.")
    st.stop()

project_name = st.session_state.get("selected_project_name", "Proyek")

# ---------- Ambil data proyek ----------
proj_rows = supabase.table("projects").select("*").eq("id", project_id).execute().data
project = proj_rows[0] if proj_rows else {"name": project_name}

rab_items = supabase.table("rab_items").select(
    "id, parent_id, volume, unit_price, is_addendum").eq("project_id", project_id).execute().data or []

periods = supabase.table("opname_periods").select("id").eq(
    "project_id", project_id).execute().data or []
period_ids = [p["id"] for p in periods]
opname_details = []
if period_ids:
    opname_details = supabase.table("opname_details").select(
        "rab_item_id, actual_volume").in_("period_id", period_ids).execute().data or []

expenses = supabase.table("project_expenses").select(
    "amount").eq("project_id", project_id).execute().data or []

payments = []
try:
    payments = supabase.table("project_payments").select("*").eq(
        "project_id", project_id).order("payment_date").execute().data or []
except Exception:
    payments = []

recap = compute_project_recap(project_id, rab_items, opname_details, expenses, payments)

# ---------- Pilihan bagian ----------
st.subheader("Pilih Bagian Laporan")
inc_header = st.checkbox("Header proyek + kop perusahaan", value=True)
inc_summary = st.checkbox("Ringkasan keuangan", value=True)
inc_payments = st.checkbox("Rincian pembayaran masuk", value=True)

# ---------- Pratinjau ringkas ----------
st.divider()
st.subheader(f"Pratinjau — {project.get('name', project_name)}")
if inc_summary:
    cc = st.columns(3)
    cc[0].metric("Nilai RAB (+Adendum)", format_rupiah(recap["nilai_rab_total"]))
    cc[1].metric("Tertagih (Opname)", format_rupiah(recap["total_opname"]))
    cc[2].metric("Uang Masuk", format_rupiah(recap["uang_masuk"]))
    cc2 = st.columns(3)
    cc2[0].metric("Piutang", format_rupiah(recap["piutang"]))
    cc2[1].metric("Pengeluaran", format_rupiah(recap["total_pengeluaran"]))
    cc2[2].metric("Margin (kas)", format_rupiah(recap["margin"]))
if inc_payments and payments:
    st.caption(f"{len(payments)} pembayaran tercatat, total {format_rupiah(recap['uang_masuk'])}.")


# ==================== GENERATOR PDF ====================
def build_pdf():
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.2 * cm, bottomMargin=1.2 * cm,
                            leftMargin=1.2 * cm, rightMargin=1.2 * cm)
    styles = getSampleStyleSheet()
    normal = styles["Normal"]
    elements = []

    if inc_header:
        try:
            build_letterhead(elements, get_company_settings(), styles)
        except Exception:
            pass

    title_style = ParagraphStyle("RepTitle", parent=styles["Normal"], fontSize=15,
                                 fontName="Helvetica-Bold", textColor=colors.HexColor("#0d6efd"),
                                 spaceAfter=6)
    elements.append(Paragraph("LAPORAN PROYEK", title_style))

    if inc_header:
        info = [
            f"<b>Proyek:</b> {project.get('name', '')}",
            f"<b>Klien:</b> {project.get('client', '') or '-'}",
            f"<b>Lokasi:</b> {project.get('location', '') or '-'}",
            f"<b>Tanggal Laporan:</b> {datetime.now().strftime('%d %B %Y')}",
        ]
        elements.append(Paragraph("<br/>".join(info), normal))
        elements.append(Spacer(1, 0.4 * cm))

    if inc_summary:
        elements.append(Paragraph("<b>Ringkasan Keuangan</b>", styles["Heading3"]))
        rows = [
            ["Nilai RAB (+ Adendum)", format_rupiah(recap["nilai_rab_total"])],
            ["Nilai Tertagih (Opname)", format_rupiah(recap["total_opname"])],
            ["Uang Masuk (Kas)", format_rupiah(recap["uang_masuk"])],
            ["Piutang (belum dibayar)", format_rupiah(recap["piutang"])],
            ["Total Pengeluaran", format_rupiah(recap["total_pengeluaran"])],
            ["Margin (kas riil)", format_rupiah(recap["margin"])],
        ]
        t = Table(rows, colWidths=[8 * cm, 6 * cm])
        t.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("BACKGROUND", (0, 5), (-1, 5), colors.HexColor("#d4edda")),
            ("FONTNAME", (0, 5), (-1, 5), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 0), (-1, 4), [colors.white, colors.HexColor("#f5f5f5")]),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 0.4 * cm))

    if inc_payments:
        elements.append(Paragraph("<b>Rincian Pembayaran Masuk</b>", styles["Heading3"]))
        if payments:
            pay_rows = [["Tanggal", "Termin", "Keterangan", "Jumlah (Rp)"]]
            for p in payments:
                pay_rows.append([
                    str(p.get("payment_date", "")),
                    p.get("termin", "") or "-",
                    Paragraph(p.get("description", "") or "-", normal),
                    f"{(p.get('amount', 0) or 0):,.0f}",
                ])
            pay_rows.append(["", "", Paragraph("<b>TOTAL</b>", normal),
                             f"{recap['uang_masuk']:,.0f}"])
            pt = Table(pay_rows, colWidths=[2.5 * cm, 3 * cm, 7 * cm, 3.5 * cm], repeatRows=1)
            pt.setStyle(TableStyle([
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d6efd")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (3, 1), (3, -1), "RIGHT"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#d4edda")),
            ]))
            elements.append(pt)
        else:
            elements.append(Paragraph("Belum ada pembayaran tercatat.", normal))

    doc.build(elements)
    buf.seek(0)
    return buf


# ==================== GENERATOR EXCEL ====================
def build_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Proyek"
    thin = Border(*[Side(style="thin")] * 4)

    base = 1
    if inc_header:
        try:
            base = add_excel_letterhead(ws, get_company_settings(), num_cols=4)
        except Exception:
            base = 1

    r = base
    ws.cell(row=r, column=1, value="LAPORAN PROYEK").font = Font(bold=True, size=14, color="0d6efd")
    r += 1
    if inc_header:
        for label, val in [("Proyek", project.get("name", "")),
                           ("Klien", project.get("client", "") or "-"),
                           ("Lokasi", project.get("location", "") or "-"),
                           ("Tanggal Laporan", datetime.now().strftime("%d %B %Y"))]:
            ws.cell(row=r, column=1, value=label).font = Font(bold=True)
            ws.cell(row=r, column=2, value=val)
            r += 1
    r += 1

    if inc_summary:
        ws.cell(row=r, column=1, value="RINGKASAN KEUANGAN").font = Font(bold=True, size=12)
        r += 1
        summary_rows = [
            ("Nilai RAB (+ Adendum)", recap["nilai_rab_total"]),
            ("Nilai Tertagih (Opname)", recap["total_opname"]),
            ("Uang Masuk (Kas)", recap["uang_masuk"]),
            ("Piutang (belum dibayar)", recap["piutang"]),
            ("Total Pengeluaran", recap["total_pengeluaran"]),
            ("Margin (kas riil)", recap["margin"]),
        ]
        for label, val in summary_rows:
            c1 = ws.cell(row=r, column=1, value=label); c1.border = thin
            c2 = ws.cell(row=r, column=2, value=val); c2.border = thin
            c2.number_format = '#,##0'
            if label.startswith("Margin"):
                c1.font = Font(bold=True); c2.font = Font(bold=True)
                c1.fill = c2.fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
            r += 1
        r += 1

    if inc_payments:
        ws.cell(row=r, column=1, value="RINCIAN PEMBAYARAN MASUK").font = Font(bold=True, size=12)
        r += 1
        headers = ["Tanggal", "Termin", "Keterangan", "Jumlah (Rp)"]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=r, column=ci, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
            c.border = thin
            c.alignment = Alignment(horizontal="center")
        r += 1
        for p in payments:
            ws.cell(row=r, column=1, value=str(p.get("payment_date", ""))).border = thin
            ws.cell(row=r, column=2, value=p.get("termin", "") or "-").border = thin
            ws.cell(row=r, column=3, value=p.get("description", "") or "-").border = thin
            cv = ws.cell(row=r, column=4, value=(p.get("amount", 0) or 0)); cv.border = thin
            cv.number_format = '#,##0'
            r += 1
        ws.cell(row=r, column=3, value="TOTAL").font = Font(bold=True)
        ct = ws.cell(row=r, column=4, value=recap["uang_masuk"]); ct.font = Font(bold=True)
        ct.number_format = '#,##0'

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 18

    out = BytesIO(); wb.save(out); out.seek(0)
    return out


# ==================== TOMBOL EXPORT ====================
st.divider()
st.subheader("Export Laporan")
ce1, ce2 = st.columns(2)
safe_name = "".join(c for c in project.get("name", "Proyek") if c.isalnum() or c in " _-").strip()[:40] or "Proyek"

with ce1:
    if st.button("📄 Buat PDF", type="primary", use_container_width=True):
        try:
            pdf = build_pdf()
            st.download_button("⬇️ Download PDF", data=pdf,
                               file_name=f"Laporan_{safe_name}_{datetime.now().strftime('%Y%m%d')}.pdf",
                               mime="application/pdf", use_container_width=True)
        except Exception as e:
            st.error(f"Gagal membuat PDF: {e}")

with ce2:
    if st.button("📊 Buat Excel", type="primary", use_container_width=True):
        try:
            xls = build_excel()
            st.download_button("⬇️ Download Excel", data=xls,
                               file_name=f"Laporan_{safe_name}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)
        except Exception as e:
            st.error(f"Gagal membuat Excel: {e}")
