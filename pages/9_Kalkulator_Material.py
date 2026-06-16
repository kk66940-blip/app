"""
pages/9_Kalkulator_Material.py
Kalkulator Kebutuhan Material Bangunan.

Port dari kalkulator React (.jsx) ke Streamlit. Input berbasis dimensi (bukan
volume jadi), hasil terhitung otomatis, plus tab Ringkasan yang menjumlahkan
material dari semua kategori menjadi daftar belanja.

Logika perhitungan ada di utils/material_calc.py (sudah diuji terpisah).
Halaman ini berdiri sendiri dan tidak menulis ke database.

Catatan: di Streamlit, setiap perubahan input otomatis menjalankan ulang
script — jadi hasil ikut ter-update tanpa tombol (mirip live update React).
"""

import sys
from io import BytesIO
from pathlib import Path
from datetime import datetime

import streamlit as st

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from utils import material_calc as mc

st.header("🧮 Kalkulator Kebutuhan Material Bangunan")
st.caption("Estimasi kebutuhan material konstruksi dari dimensi pekerjaan.")

st.warning(
    "⚠️ Koefisien adalah perkiraan untuk perencanaan awal (RAB), sebagian "
    "angka praktis lapangan — bukan semuanya SNI resmi. Tambahkan margin untuk "
    "sisa/waste dan sesuaikan kondisi lapangan. Untuk besi tulangan & rangka "
    "atap, konfirmasikan ke tenaga ahli/aplikator."
)

# Simpan hasil tiap kategori untuk Ringkasan
if "matcalc" not in st.session_state:
    st.session_state.matcalc = {}


def _row(label, value, dec, unit, note=None):
    """Render satu baris hasil."""
    c1, c2 = st.columns([3, 2])
    c1.markdown(f"{label}")
    txt = f"**{mc.fmt(value, dec)}** {unit}"
    if note:
        txt += f"  \n<span style='color:#d97706;font-size:0.85em'>{note}</span>"
    c2.markdown(txt, unsafe_allow_html=True)


tabs = st.tabs([
    "Pondasi", "Beton", "Dinding", "Plester",
    "Lantai", "Atap", "Plafon", "Cat", "📋 Ringkasan",
])

# ============================ PONDASI ============================
with tabs[0]:
    st.subheader("🏠 Pondasi")
    st.caption("Pasangan batu kali, campuran 1:5")
    panjang = st.number_input("Panjang total pondasi (m)", min_value=0.0, value=24.0, step=0.5, key="pd_p")
    c1, c2 = st.columns(2)
    lebar_atas = c1.number_input("Lebar atas (m)", min_value=0.0, value=0.3, step=0.05, key="pd_la")
    lebar_bawah = c2.number_input("Lebar bawah (m)", min_value=0.0, value=0.6, step=0.05, key="pd_lb")
    tinggi = st.number_input("Tinggi pondasi (m)", min_value=0.0, value=0.7, step=0.05, key="pd_t")

    r = mc.calc_pondasi(panjang, lebar_atas, lebar_bawah, tinggi)
    st.session_state.matcalc["pondasi"] = r
    st.markdown("##### Kebutuhan Material")
    _row("Volume pasangan", r["volume"], 2, "m³")
    _row("Batu kali", r["batu_kali_m3"], 2, "m³")
    import math
    _row("Semen", r["semen_kg"], 0, "kg", f"≈ {math.ceil(r['semen_kg']/50)} sak (50kg)")
    _row("Pasir pasang", r["pasir_m3"], 2, "m³")
    st.info("Volume dari penampang trapesium (rata-rata lebar atas & bawah) × tinggi × panjang jalur pondasi.")

# ============================ BETON ============================
with tabs[1]:
    import math
    st.subheader("📦 Beton Bertulang")
    st.caption("Untuk sloof, kolom, atau balok")
    c1, c2 = st.columns(2)
    b_panjang = c1.number_input("Panjang (m)", min_value=0.0, value=3.0, step=0.1, key="bt_p")
    b_lebar = c2.number_input("Lebar (m)", min_value=0.0, value=0.15, step=0.01, key="bt_l")
    c3, c4 = st.columns(2)
    b_tinggi = c3.number_input("Tinggi / Tebal (m)", min_value=0.0, value=0.15, step=0.01, key="bt_t")
    b_jumlah = c4.number_input("Jumlah (bh)", min_value=0, value=8, step=1, key="bt_j")
    b_rasio = st.selectbox(
        "Rasio campuran (semen : pasir : kerikil)",
        list(mc.RASIO_BETON.keys()),
        format_func=lambda x: {
            "1:2:3": "1 : 2 : 3 — Umum (sloof/kolom/balok)",
            "1:1.5:2.5": "1 : 1,5 : 2,5 — Mutu tinggi",
            "1:3:5": "1 : 3 : 5 — Rabat / lantai kerja",
        }[x],
        key="bt_r",
    )

    st.markdown("**Tulangan Besi**")
    c5, c6 = st.columns(2)
    b_jml_utama = c5.number_input("Jumlah tulangan utama (btg)", min_value=0, value=4, step=1, key="bt_ju")
    b_dia_utama = c6.selectbox("Diameter utama", [10, 12, 13, 16, 19], index=1,
                               format_func=lambda d: f"D{d}", key="bt_du")
    c7, c8 = st.columns(2)
    b_dia_seng = c7.selectbox("Diameter sengkang", [6, 8, 10], index=1,
                              format_func=lambda d: f"Ø{d}", key="bt_ds")
    b_jarak_seng = c8.number_input("Jarak sengkang (cm)", min_value=0.0, value=15.0, step=1.0, key="bt_js")

    r = mc.calc_beton(b_panjang, b_lebar, b_tinggi, b_jumlah, b_rasio,
                      b_jml_utama, b_dia_utama, b_dia_seng, b_jarak_seng)
    st.session_state.matcalc["beton"] = r

    st.markdown("##### Adukan Beton")
    _row("Volume beton", r["volume"], 2, "m³")
    _row("Semen", r["semen_kg"], 0, "kg", f"≈ {math.ceil(r['semen_kg']/50)} sak (50kg)" if r["semen_kg"] > 0 else None)
    _row("Pasir beton", r["pasir_m3"], 2, "m³")
    _row("Kerikil / split", r["kerikil_m3"], 2, "m³")
    st.markdown("##### Besi Tulangan")
    _row(f"Tulangan utama D{b_dia_utama}", r["berat_utama_kg"], 0, "kg",
         f"{b_jml_utama} batang ≈ {r['lonjor_utama']} lonjor (12m)")
    _row(f"Sengkang/begel Ø{b_dia_seng}", r["berat_sengkang_kg"], 0, "kg",
         f"{r['total_sengkang']} buah ≈ {r['lonjor_sengkang']} lonjor (12m)")
    _row("Total besi", r["total_besi_kg"], 0, "kg")
    st.info("Adukan pakai faktor susut 1,5×. Sengkang dari keliling penampang × jumlah (panjang ÷ jarak). Besi per lonjor 12m. Untuk struktur penting, pakai hitungan gambar kerja/insinyur.")

# ============================ DINDING ============================
with tabs[2]:
    st.subheader("🧱 Dinding")
    st.caption("Dinding ½ bata, campuran 1:5")
    c1, c2 = st.columns(2)
    d_panjang = c1.number_input("Panjang dinding (m)", min_value=0.0, value=12.0, step=0.5, key="dd_p")
    d_tinggi = c2.number_input("Tinggi dinding (m)", min_value=0.0, value=3.0, step=0.1, key="dd_t")
    d_bukaan = st.number_input("Luas bukaan (pintu + jendela) (m²)", min_value=0.0, value=4.0, step=0.5, key="dd_b")
    d_jenis = st.selectbox("Jenis dinding", ["merah", "batako", "hebel"],
                           format_func=lambda x: {"merah": "Bata Merah", "batako": "Batako", "hebel": "Bata Ringan (Hebel)"}[x],
                           key="dd_j")

    r = mc.calc_dinding(d_panjang, d_tinggi, d_bukaan, d_jenis)
    st.session_state.matcalc["dinding"] = r
    st.markdown("##### Kebutuhan Material")
    _row("Luas dinding bersih", r["luas_bersih"], 2, "m²")
    _row(r["nama_bata"], r["jumlah_bata"], 0, "buah")
    if d_jenis != "hebel":
        import math
        _row("Semen pasang", r["semen_kg"], 0, "kg", f"≈ {math.ceil(r['semen_kg']/50)} sak (50kg)" if r["semen_kg"] > 0 else None)
        _row("Pasir pasang", r["pasir_m3"], 2, "m³")
    else:
        import math
        _row("Mortar instan", r["mortar_kg"], 0, "kg", f"≈ {math.ceil(r['mortar_kg']/40)} sak (40kg)" if r["mortar_kg"] > 0 else None)
    st.info("Luas bersih = (panjang × tinggi) − bukaan. Belum termasuk plesteran (lihat tab Plester).")

# ============================ PLESTER ============================
with tabs[3]:
    import math
    st.subheader("⬜ Plesteran")
    st.caption("Plester + acian (finishing halus)")
    pl_luas = st.number_input("Luas yang diplester (m²)", min_value=0.0, value=24.0, step=0.5, key="pl_l")
    pl_tebal = st.number_input("Tebal plesteran (cm)", min_value=0.0, value=1.5, step=0.5, key="pl_t")
    pl_jenis = st.selectbox("Campuran", ["1:4", "1:3"],
                            format_func=lambda x: {"1:4": "1 : 4 — Plester biasa", "1:3": "1 : 3 — Trasram/kedap air (KM)"}[x],
                            key="pl_j")
    pl_acian = st.checkbox("Sertakan acian (finishing halus)", value=True, key="pl_a")

    r = mc.calc_plesteran(pl_luas, pl_tebal, pl_jenis, pl_acian)
    st.session_state.matcalc["plesteran"] = r
    st.markdown("##### Kebutuhan Material")
    _row("Volume adukan", r["volume"], 2, "m³")
    _row("Semen (plester + acian)", r["semen_kg"], 0, "kg", f"≈ {math.ceil(r['semen_kg']/50)} sak (50kg)" if r["semen_kg"] > 0 else None)
    _row("Pasir", r["pasir_m3"], 2, "m³")
    st.info("Luas plesteran umumnya ±2× luas dinding (dua sisi). Acian ±2,5 kg semen/m², sudah termasuk total semen di atas.")

# ============================ LANTAI ============================
with tabs[4]:
    import math
    st.subheader("⬛ Lantai & Keramik")
    st.caption("Sudah termasuk +7% cadangan potongan")
    l_luas = st.number_input("Luas lantai (m²)", min_value=0.0, value=36.0, step=0.5, key="ln_l")
    l_ukuran = st.selectbox("Ukuran keramik", ["30", "40", "50", "60", "80", "custom"],
                            index=3, format_func=lambda x: "Ukuran custom" if x == "custom" else f"{x} × {x} cm",
                            key="ln_u")
    l_cp, l_cl = 40, 40
    if l_ukuran == "custom":
        cc1, cc2 = st.columns(2)
        l_cp = cc1.number_input("Panjang (cm)", min_value=1.0, value=40.0, step=1.0, key="ln_cp")
        l_cl = cc2.number_input("Lebar (cm)", min_value=1.0, value=40.0, step=1.0, key="ln_cl")
    l_isidus = st.number_input("Isi per dus (opsional)", min_value=0, value=0, step=1, key="ln_id")

    r = mc.calc_lantai(l_luas, l_ukuran, l_cp, l_cl, l_isidus)
    st.session_state.matcalc["lantai"] = r
    st.markdown("##### Kebutuhan Material")
    _row("Keramik", r["jumlah_keping"], 0, "keping", f"≈ {r['jumlah_dus']} dus" if r["jumlah_dus"] > 0 else None)
    _row("Semen instan / perekat", r["semen_instan_kg"], 0, "kg", f"≈ {math.ceil(r['semen_instan_kg']/40)} sak (40kg)" if r["semen_instan_kg"] > 0 else None)
    _row("Nat keramik", r["nat_kg"], 0, "kg")
    st.info("Jumlah keping +7% cadangan potongan. Kebutuhan nat bervariasi sesuai ukuran keramik & lebar nat.")

# ============================ ATAP ============================
with tabs[5]:
    st.subheader("🔺 Atap")
    st.caption("Gunakan luas atap sisi miring, bukan luas bangunan")
    a_luas = st.number_input("Luas atap (sisi miring) (m²)", min_value=0.0, value=90.0, step=1.0, key="at_l")
    a_jenis = st.selectbox("Jenis penutup atap", ["keramik", "beton", "metal"],
                           format_func=lambda x: {"keramik": "Genteng Keramik", "beton": "Genteng Beton", "metal": "Genteng Metal Berpasir"}[x],
                           key="at_j")
    a_rangka = st.checkbox("Hitung perkiraan rangka baja ringan", value=False, key="at_r")

    r = mc.calc_atap(a_luas, a_jenis, a_rangka)
    st.session_state.matcalc["atap"] = r
    st.markdown("##### Kebutuhan Material")
    _row(r["nama_genteng"], r["jumlah_genteng"], 0, "buah", "+5% cadangan")
    if a_rangka:
        _row("Rangka baja ringan", r["baja_ringan_batang"], 0, "batang (6m)")
    st.info("Jika hanya tahu luas denah, kalikan ±1,2–1,3 untuk perkiraan luas atap. Kebutuhan rangka kasar — konfirmasi ke aplikator.")

# ============================ PLAFON ============================
with tabs[6]:
    st.subheader("▦ Plafon")
    st.caption("Rangka hollow + wall angle + drop ceiling")
    c1, c2 = st.columns(2)
    pf_luas = c1.number_input("Luas plafon (m²)", min_value=0.0, value=45.0, step=0.5, key="pf_l")
    pf_kel = c2.number_input("Keliling ruangan (m)", min_value=0.0, value=27.0, step=0.5, key="pf_k")
    pf_jenis = st.selectbox("Jenis panel", ["gypsum", "grc"],
                            format_func=lambda x: {"gypsum": "Gypsum 9mm (120×240cm)", "grc": "GRC / Kalsiboard (122×244cm)"}[x],
                            key="pf_j")
    pf_drop = st.checkbox("Pakai plafon drop / indirect", value=False, key="pf_d")
    pf_pdrop, pf_ldrop = 10.0, 20.0
    if pf_drop:
        cc1, cc2 = st.columns(2)
        pf_pdrop = cc1.number_input("Panjang dropan (m)", min_value=0.0, value=10.0, step=0.5, key="pf_pd")
        pf_ldrop = cc2.number_input("Lebar dropan (cm)", min_value=0.0, value=20.0, step=1.0, key="pf_ld")

    r = mc.calc_plafon(pf_luas, pf_kel, pf_jenis, pf_drop, pf_pdrop, pf_ldrop)
    st.session_state.matcalc["plafon"] = r
    st.markdown("##### Kebutuhan Material")
    _row(r["nama_plafon"], r["jumlah_lembar"], 0, "lembar", "+5% cadangan")
    _row("Rangka atas (penggantung)", r["rangka_atas_batang"], 0, "batang (4m)")
    _row("Rangka bawah (dudukan)", r["rangka_bawah_batang"], 0, "batang (4m)")
    _row("Wall angle", r["wall_angle_batang"], 0, "batang (3m)")
    if pf_drop:
        _row("Hollow dropan", r["dropan_batang"], 0, "batang (4m)")
    _row("Sekrup", r["sekrup"], 0, "buah")
    st.info("Rangka atas ±1,6 m/m² & bawah ±2 m/m². Wall angle dari keliling ruangan (batang 3m). Plafon drop menambah luas panel & hollow.")

# ============================ CAT ============================
with tabs[7]:
    st.subheader("🪣 Cat")
    st.caption("Cat tembok / dinding")
    c_luas = st.number_input("Luas yang dicat (m²)", min_value=0.0, value=150.0, step=1.0, key="ct_l")
    c1, c2 = st.columns(2)
    c_lapis = c1.number_input("Jumlah lapis", min_value=0, value=2, step=1, key="ct_lp")
    c_daya = c2.number_input("Daya sebar (m²/L)", min_value=0.1, value=10.0, step=0.5, key="ct_d")
    c_kemasan = st.selectbox("Ukuran kemasan", [2.5, 5.0, 20.0],
                             format_func=lambda x: f"{mc.fmt(x,1)} Liter / kg", key="ct_k")

    r = mc.calc_cat(c_luas, c_lapis, c_daya, c_kemasan)
    st.session_state.matcalc["cat"] = r
    st.markdown("##### Kebutuhan Material")
    _row("Total cat", r["total_liter"], 1, "liter")
    _row("Jumlah kemasan", r["jumlah_kemasan"], 0, f"× {mc.fmt(c_kemasan,1)} L")
    st.info("Daya sebar standar ±10 m²/liter per lapis (bervariasi per produk). Total sudah memperhitungkan jumlah lapis.")

# ============================ RINGKASAN ============================
with tabs[8]:
    import math
    st.subheader("📋 Ringkasan — Daftar Belanja Material")
    st.caption("Total gabungan dari semua kategori yang sudah kamu isi.")

    d = st.session_state.matcalc
    pdn, btn, ddn = d.get("pondasi", {}), d.get("beton", {}), d.get("dinding", {})
    pln, lnn, atn = d.get("plesteran", {}), d.get("lantai", {}), d.get("atap", {})
    pfn, ctn = d.get("plafon", {}), d.get("cat", {})

    semen_pc = (pdn.get("semen_kg", 0) + btn.get("semen_kg", 0)
                + (ddn.get("semen_kg", 0) if ddn.get("jenis") != "hebel" else 0)
                + pln.get("semen_kg", 0))
    pasir = (pdn.get("pasir_m3", 0) + btn.get("pasir_m3", 0)
             + (ddn.get("pasir_m3", 0) if ddn.get("jenis") != "hebel" else 0)
             + pln.get("pasir_m3", 0))
    mortar = ((ddn.get("mortar_kg", 0) if ddn.get("jenis") == "hebel" else 0)
              + lnn.get("semen_instan_kg", 0))
    hollow = (pfn.get("rangka_atas_batang", 0) + pfn.get("rangka_bawah_batang", 0)
              + pfn.get("dropan_batang", 0))

    items = [
        ("Semen PC (Portland)", semen_pc, 0, "kg", f"≈ {math.ceil(semen_pc/50)} sak (50kg)" if semen_pc > 0 else None),
        ("Pasir", pasir, 2, "m³", None),
        ("Batu kali", pdn.get("batu_kali_m3", 0), 2, "m³", None),
        ("Kerikil / split", btn.get("kerikil_m3", 0), 2, "m³", None),
        ("Besi tulangan", btn.get("total_besi_kg", 0), 0, "kg", None),
        ("Mortar instan", mortar, 0, "kg", f"≈ {math.ceil(mortar/40)} sak (40kg)" if mortar > 0 else None),
        (ddn.get("nama_bata", "Bata"), ddn.get("jumlah_bata", 0), 0, "buah", None),
        (f"Keramik ({lnn.get('ukuran_label','-')})", lnn.get("jumlah_keping", 0), 0, "keping",
         f"≈ {lnn.get('jumlah_dus',0)} dus" if lnn.get("jumlah_dus", 0) > 0 else None),
        ("Nat keramik", lnn.get("nat_kg", 0), 0, "kg", None),
        (atn.get("nama_genteng", "Genteng"), atn.get("jumlah_genteng", 0), 0, "buah", "+5% cadangan"),
        ("Rangka baja ringan", atn.get("baja_ringan_batang", 0), 0, "batang", None),
        (pfn.get("nama_plafon", "Plafon"), pfn.get("jumlah_lembar", 0), 0, "lembar", "+5% cadangan"),
        ("Hollow plafon (total)", hollow, 0, "batang (4m)", None),
        ("Wall angle plafon", pfn.get("wall_angle_batang", 0), 0, "batang (3m)", None),
        ("Cat", ctn.get("total_liter", 0), 1, "liter",
         f"≈ {ctn.get('jumlah_kemasan',0)} kemasan" if ctn.get("jumlah_kemasan", 0) > 0 else None),
    ]
    items = [it for it in items if (it[1] or 0) > 0]

    if not items:
        st.info("Belum ada data. Isi dulu tab kategori di atas.")
    else:
        for label, value, dec, unit, note in items:
            _row(label, value, dec, unit, note)

        # Export Excel
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Daftar Belanja"
        thin = Border(*[Side(style="thin")] * 4)
        ws.merge_cells("A1:C1")
        ws["A1"] = "DAFTAR BELANJA MATERIAL"
        ws["A1"].font = Font(bold=True, size=14, color="0d6efd")
        ws["A2"] = f"Tanggal: {datetime.now().strftime('%d %B %Y')}"
        for ci, h in enumerate(["Material", "Jumlah", "Satuan"], 1):
            cell = ws.cell(row=4, column=ci, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin
        rr = 5
        for label, value, dec, unit, note in items:
            ws.cell(row=rr, column=1, value=label).border = thin
            ws.cell(row=rr, column=2, value=round(value, dec)).border = thin
            ws.cell(row=rr, column=3, value=unit).border = thin
            rr += 1
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 14
        out = BytesIO(); wb.save(out); out.seek(0)

        st.download_button(
            "⬇️ Download Daftar Belanja (Excel)", data=out,
            file_name=f"Material_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    st.info("Semua angka perkiraan untuk perencanaan awal (RAB). Tambah margin waste & sesuaikan harga wilayahmu. Besi & rangka atap konfirmasi ke ahli.")
