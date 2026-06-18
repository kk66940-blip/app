"""
utils/export_utils.py
Versi lengkap + sudah diperbaiki untuk mendukung RAP hierarchy
"""

from io import BytesIO
from datetime import datetime
from collections import defaultdict
from typing import List, Dict, Any, Optional, Callable

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import cm


# ==================== STYLING ====================
HEADER_FILL = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
SECTION_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBTOTAL_FILL = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
SUBTOTAL_FONT = Font(bold=True, size=10)
GRAND_FILL = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
GRAND_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


def _build_children_map(items: List[Dict], id_key: str = "id", parent_key: str = "parent_id") -> Dict:
    children_map = defaultdict(list)
    for item in items:
        parent = item.get(parent_key)
        children_map[parent].append(item)
    
    for pid in children_map:
        children_map[pid] = sorted(children_map[pid], key=lambda x: (x.get('sort_order', 0), x.get(id_key, 0)))
    return children_map


def _get_root_items(items: List[Dict], id_key: str = "id", parent_key: str = "parent_id") -> List[Dict]:
    all_ids = {item.get(id_key) for item in items if item.get(id_key)}
    roots = [item for item in items if item.get(parent_key) is None or item.get(parent_key) not in all_ids]
    return sorted(roots, key=lambda x: (x.get('sort_order', 0), x.get(id_key, 0)))


# ==================== EXCEL EXPORT (LENGKAP) ====================
def export_hierarchical_excel(
    items: List[Dict],
    project_name: str,
    title: str = "RENCANA ANGGARAN PELAKSANAAN (RAP)",
    filename_prefix: str = "RAP",
    get_total_func: Optional[Callable[[Dict], float]] = None,
    id_key: str = "id",
    parent_key: str = "parent_id"
) -> BytesIO:
    if not items:
        raise ValueError("No data to export")

    children_map = _build_children_map(items, id_key=id_key, parent_key=parent_key)
    root_items = _get_root_items(items, id_key=id_key, parent_key=parent_key)

    # Bobot pekerjaan (%) terhadap grand total dari nilai item daun.
    def _val(it):
        v = (it.get('volume', 0) or 0)
        p = (it.get('unit_price', 0) or it.get('execution_price', 0) or 0)
        return get_total_func(it) if get_total_func else (v * p)
    _grand_for_weight = sum(_val(it) for it in items if not children_map.get(it.get(id_key)))
    def _weight(it):
        return (_val(it) / _grand_for_weight * 100.0) if _grand_for_weight > 0 else 0.0

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename_prefix[:31]

    # Kop surat perusahaan (gagal-aman). Mengembalikan baris awal konten.
    base = 1
    try:
        from utils.company import get_company_settings, add_excel_letterhead
        base = add_excel_letterhead(ws, get_company_settings(), num_cols=7)
    except Exception:
        base = 1

    title_row = base          # judul
    date_row = base + 1       # tanggal
    header_row = base + 3     # header tabel (1 baris kosong pemisah)

    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=7)
    tcell = ws.cell(row=title_row, column=1, value=f"{title} - {project_name}")
    tcell.font = Font(bold=True, size=14, color="0d6efd")
    tcell.alignment = Alignment(horizontal='center')

    ws.merge_cells(start_row=date_row, start_column=1, end_row=date_row, end_column=7)
    dcell = ws.cell(row=date_row, column=1, value=f"Tanggal Export: {datetime.now().strftime('%d %B %Y')}")
    dcell.font = Font(italic=True, size=10)

    headers = ["No", "Uraian Pekerjaan", "Satuan", "Volume", "Harga Satuan (Rp)", "Jumlah (Rp)", "Bobot (%)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    row_num = header_row + 1
    grand_total = 0

    def process_node(node: Dict, path: List[int]) -> float:
        nonlocal row_num, grand_total
        node_id = node.get(id_key)
        children = children_map.get(node_id, [])
        has_children = len(children) > 0
        current_no = ".".join(map(str, path))

        if has_children:
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)
            cell = ws.cell(row=row_num, column=1, value=f"▶ {node.get('description', '')}")
            cell.font = SECTION_FONT
            cell.fill = SECTION_FILL
            for c in range(1, 8):
                ws.cell(row=row_num, column=c).border = THIN_BORDER
                ws.cell(row=row_num, column=c).fill = SECTION_FILL
            row_num += 1

            sub_total = 0
            for idx, child in enumerate(children, 1):
                sub_total += process_node(child, path + [idx])
            
            ws.cell(row=row_num, column=2, value="SUBTOTAL").font = SUBTOTAL_FONT
            ws.cell(row=row_num, column=6, value=sub_total).font = SUBTOTAL_FONT
            ws.cell(row=row_num, column=6).number_format = '#,##0'
            # Bobot grup = subtotal / grand total (rollup dari anak-anaknya)
            _grp_w = (sub_total / _grand_for_weight * 100.0) if _grand_for_weight > 0 else 0.0
            ws.cell(row=row_num, column=7, value=_grp_w).font = SUBTOTAL_FONT
            ws.cell(row=row_num, column=7).number_format = '0.00"%"'
            for c in range(1, 8):
                ws.cell(row=row_num, column=c).border = THIN_BORDER
                ws.cell(row=row_num, column=c).fill = SUBTOTAL_FILL
            row_num += 1
            return sub_total
        else:
            vol = node.get('volume', 0) or 0
            price = node.get('unit_price', 0) or node.get('execution_price', 0) or 0
            total = get_total_func(node) if get_total_func else (vol * price)
            grand_total += total

            ws.cell(row=row_num, column=1, value=current_no).border = THIN_BORDER
            ws.cell(row=row_num, column=2, value=f"    {node.get('description', '')}").border = THIN_BORDER
            ws.cell(row=row_num, column=3, value=node.get('unit', '')).border = THIN_BORDER
            ws.cell(row=row_num, column=4, value=vol).border = THIN_BORDER
            ws.cell(row=row_num, column=5, value=price).border = THIN_BORDER
            ws.cell(row=row_num, column=6, value=total).border = THIN_BORDER
            ws.cell(row=row_num, column=7, value=_weight(node)).border = THIN_BORDER

            ws.cell(row=row_num, column=4).number_format = '#,##0.00'
            ws.cell(row=row_num, column=5).number_format = '#,##0'
            ws.cell(row=row_num, column=6).number_format = '#,##0'
            ws.cell(row=row_num, column=7).number_format = '0.00"%"'
            row_num += 1
            return total

    for idx, root in enumerate(root_items, 1):
        process_node(root, [idx])

    # Grand Total
    row_num += 1
    ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
    ws.cell(row=row_num, column=2, value="GRAND TOTAL").font = GRAND_FONT
    ws.cell(row=row_num, column=2).fill = GRAND_FILL
    ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='right')

    ws.cell(row=row_num, column=6, value=grand_total).font = GRAND_FONT
    ws.cell(row=row_num, column=6).fill = GRAND_FILL
    ws.cell(row=row_num, column=6).number_format = '#,##0'
    # Bobot total = 100% (bila ada nilai)
    ws.cell(row=row_num, column=7, value=(100.0 if _grand_for_weight > 0 else 0.0)).font = GRAND_FONT
    ws.cell(row=row_num, column=7).fill = GRAND_FILL
    ws.cell(row=row_num, column=7).number_format = '0.00"%"'
    for c in range(2, 8):
        ws.cell(row=row_num, column=c).border = THIN_BORDER
        ws.cell(row=row_num, column=c).fill = GRAND_FILL

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 12

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ==================== PDF EXPORT (DIPERBAIKI) ====================
def export_hierarchical_pdf(
    items: List[Dict],
    project_name: str,
    title: str = "RENCANA ANGGARAN PELAKSANAAN (RAP)",
    filename_prefix: str = "RAP",
    get_total_func: Optional[Callable[[Dict], float]] = None,
    id_key: str = "id",
    parent_key: str = "parent_id"
) -> BytesIO:
    if not items:
        raise ValueError("No data to export")

    children_map = _build_children_map(items, id_key=id_key, parent_key=parent_key)
    root_items = _get_root_items(items, id_key=id_key, parent_key=parent_key)

    # Bobot pekerjaan (%) terhadap grand total nilai item daun.
    def _val_pdf(it):
        v = (it.get('volume', 0) or 0)
        p = (it.get('unit_price', 0) or it.get('execution_price', 0) or 0)
        return get_total_func(it) if get_total_func else (v * p)
    _grand_weight_pdf = sum(_val_pdf(it) for it in items if not children_map.get(it.get(id_key)))
    def _weight_pdf(it):
        return (_val_pdf(it) / _grand_weight_pdf * 100.0) if _grand_weight_pdf > 0 else 0.0

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=1.2*cm, leftMargin=1.2*cm,
                            topMargin=1.2*cm, bottomMargin=1.2*cm)

    styles = getSampleStyleSheet()
    normal = ParagraphStyle('Normal', parent=styles['Normal'], fontSize=8, leading=10)

    elements = []
    # Kop surat perusahaan (gagal-aman bila helper/data tak tersedia).
    try:
        from utils.company import get_company_settings, build_letterhead
        build_letterhead(elements, get_company_settings(), styles)
    except Exception:
        pass
    elements.append(Paragraph(f"<b>{title}</b>", ParagraphStyle('Title', fontSize=14, fontName='Helvetica-Bold', textColor=colors.HexColor('#0d6efd'))))
    elements.append(Paragraph(f"<b>Proyek:</b> {project_name}", normal))
    elements.append(Paragraph(f"<b>Tanggal:</b> {datetime.now().strftime('%d %B %Y')}", normal))
    elements.append(Spacer(1, 0.3*cm))

    table_data = [["No", "Uraian Pekerjaan", "Sat", "Vol", "Harga (Rp)", "Jumlah (Rp)", "Bobot"]]
    header_row_indices, subtotal_row_indices = [], []
    grand_total = 0
    current_row = 1

    def add_to_pdf(node: Dict, path: List[int]) -> float:
        nonlocal grand_total, current_row
        node_id = node.get(id_key)
        children = children_map.get(node_id, [])
        has_children = len(children) > 0
        current_no = ".".join(map(str, path))

        if has_children:
            header_row_indices.append(current_row)
            table_data.append(["", Paragraph(f"▶ {node.get('description', '')}", 
                               ParagraphStyle('Header', fontSize=9, fontName='Helvetica-Bold', textColor=colors.white)), "", "", "", "", ""])
            current_row += 1

            sub_total = 0
            for idx, child in enumerate(children, 1):
                sub_total += add_to_pdf(child, path + [idx])
            
            subtotal_row_indices.append(current_row)
            _grp_w_pdf = (sub_total / _grand_weight_pdf * 100.0) if _grand_weight_pdf > 0 else 0.0
            table_data.append(["", Paragraph("<b>SUBTOTAL</b>", normal), "", "", "", Paragraph(f"<b>{sub_total:,.0f}</b>", normal),
                               Paragraph(f"<b>{_grp_w_pdf:.2f}%</b>", normal)])
            current_row += 1
            return sub_total
        else:
            vol = node.get('volume', 0) or 0
            price = node.get('unit_price', 0) or node.get('execution_price', 0) or 0
            total = get_total_func(node) if get_total_func else (vol * price)
            grand_total += total

            table_data.append([
                current_no,
                Paragraph(f"    {node.get('description', '')}", normal),
                node.get('unit', ''),
                f"{vol:,.2f}",
                f"{price:,.0f}",
                f"{total:,.0f}",
                f"{_weight_pdf(node):.2f}%"
            ])
            current_row += 1
            return total

    for idx, root in enumerate(root_items, 1):
        add_to_pdf(root, [idx])

    # Grand Total
    table_data.append(["", Paragraph("<b>GRAND TOTAL</b>", normal), "", "", "", Paragraph(f"<b>{grand_total:,.0f}</b>", normal),
                       Paragraph(f"<b>{100.0 if _grand_weight_pdf > 0 else 0.0:.2f}%</b>", normal)])

    col_widths = [1.0*cm, 7.4*cm, 1.0*cm, 1.6*cm, 2.5*cm, 2.7*cm, 1.5*cm]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)

    style_commands = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('FONTSIZE', (0, 1), (-1, -1), 7.5),
        ('ALIGN', (3, 1), (6, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]

    for row_idx in header_row_indices:
        style_commands.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#28a745')))
        style_commands.append(('TEXTCOLOR', (0, row_idx), (-1, row_idx), colors.white))
        style_commands.append(('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold'))

    for row_idx in subtotal_row_indices:
        style_commands.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#fff3cd')))
        style_commands.append(('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold'))

    grand_row = len(table_data) - 1
    style_commands.append(('BACKGROUND', (0, grand_row), (-1, grand_row), colors.HexColor('#d4edda')))
    style_commands.append(('FONTNAME', (0, grand_row), (-1, grand_row), 'Helvetica-Bold'))
    style_commands.append(('FONTSIZE', (0, grand_row), (-1, grand_row), 9))

    t.setStyle(TableStyle(style_commands))
    elements.append(t)

    doc.build(elements)
    buffer.seek(0)
    return buffer


# ==================== CONVENIENCE WRAPPERS (UNTUK KOMPATIBILITAS) ====================

def export_rab_excel(items: List[Dict], project_name: str) -> BytesIO:
    """Wrapper khusus untuk halaman RAB"""
    return export_hierarchical_excel(
        items=items,
        project_name=project_name,
        title="RENCANA ANGGARAN BIAYA (RAB)",
        filename_prefix="RAB"
    )

def export_rab_pdf(items: List[Dict], project_name: str) -> BytesIO:
    """Wrapper khusus untuk halaman RAB"""
    return export_hierarchical_pdf(
        items=items,
        project_name=project_name,
        title="RENCANA ANGGARAN BIAYA (RAB)",
        filename_prefix="RAB"
    )

def export_rap_excel(items: List[Dict], project_name: str) -> BytesIO:
    """Wrapper khusus untuk halaman RAP"""
    return export_hierarchical_excel(
        items=items,
        project_name=project_name,
        title="RENCANA ANGGARAN PELAKSANAAN (RAP)",
        filename_prefix="RAP",
        id_key="rab_item_id",
        parent_key="parent_id"
    )

def export_rap_pdf(items: List[Dict], project_name: str) -> BytesIO:
    """Wrapper khusus untuk halaman RAP"""
    return export_hierarchical_pdf(
        items=items,
        project_name=project_name,
        title="RENCANA ANGGARAN PELAKSANAAN (RAP)",
        filename_prefix="RAP",
        id_key="rab_item_id",
        parent_key="parent_id"
    )