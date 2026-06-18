"""
utils/company.py
Helper terpusat untuk data perusahaan (tabel company_settings) dan pembuatan
kop surat (letterhead) untuk dokumen PDF (invoice & SPK).

Tabel company_settings bersifat global (satu baris untuk seluruh aplikasi),
bukan per-proyek. Pola: ambil baris pertama; jika belum ada, kembalikan default.
"""

from io import BytesIO
from pathlib import Path

from utils.supabase_client import get_supabase

# Path logo bawaan yang dibundel di app (assets/logo_kmu.png).
_BUNDLED_LOGO = str(Path(__file__).resolve().parent.parent / "assets" / "logo_kmu.png")

# Default aman bila tabel belum diisi (diisi data CV Kreasindo Mandiri Utama).
_DEFAULTS = {
    "company_name": "CV KREASINDO MANDIRI UTAMA",
    "address": "JL KOPO SUKALEEUR NO 9 RT03/02, KEL. BABAKAN ASIH, "
               "KEC. BOJONGLOA KALER, KOTA BANDUNG",
    "phone": "+62 823 4346 2021",
    "email": "kreasindomandiriutama.25@gmail.com",
    "npwp": "",
    "logo_path": _BUNDLED_LOGO,
    "footer": "",
    "bank_name": "",
    "account_number": "",
    "account_holder": "",
}


def get_company_settings() -> dict:
    """Ambil data perusahaan (baris pertama). Kembalikan default bila kosong."""
    try:
        supabase = get_supabase()
        res = supabase.table("company_settings").select("*").limit(1).execute()
        if res.data:
            # Gabungkan dengan default agar key yang null tetap ada.
            merged = {**_DEFAULTS, **{k: v for k, v in res.data[0].items() if v is not None}}
            merged["id"] = res.data[0].get("id")
            return merged
    except Exception:
        pass
    return {**_DEFAULTS, "id": None}


def save_company_settings(data: dict) -> bool:
    """Simpan data perusahaan (update bila ada baris, insert bila belum)."""
    supabase = get_supabase()
    payload = {k: data.get(k, "") for k in _DEFAULTS}
    existing = supabase.table("company_settings").select("id").limit(1).execute().data
    if existing:
        supabase.table("company_settings").update(payload).eq("id", existing[0]["id"]).execute()
    else:
        supabase.table("company_settings").insert(payload).execute()
    return True


def _try_load_logo(logo_path: str):
    """Coba muat logo (dari URL ATAU file lokal) untuk reportlab Image.

    Kembalikan Image atau None. Aman gagal: jika path kosong/tidak valid/gagal
    dimuat, kembalikan None agar PDF tetap dibuat tanpa logo.
    """
    if not logo_path:
        return None
    try:
        from reportlab.platypus import Image
        from reportlab.lib.units import cm

        path_str = str(logo_path)
        if path_str.startswith(("http://", "https://")):
            import urllib.request
            with urllib.request.urlopen(path_str, timeout=5) as resp:
                src = BytesIO(resp.read())
        else:
            # File lokal (mis. logo bundel di assets/)
            import os
            if not os.path.exists(path_str):
                return None
            src = path_str

        img = Image(src)
        # Skala ke tinggi maksimal ~1.8cm menjaga rasio.
        max_h = 1.8 * cm
        if img.imageHeight > 0:
            ratio = img.imageWidth / img.imageHeight
            img.drawHeight = max_h
            img.drawWidth = max_h * ratio
        return img
    except Exception:
        return None


def build_letterhead(elements: list, company: dict, styles) -> None:
    """Tambahkan blok kop surat ke daftar `elements` reportlab (in-place).

    Parameters
    ----------
    elements : list   daftar flowable reportlab yang sedang dibangun.
    company  : dict   hasil get_company_settings().
    styles   : dict-like  hasil getSampleStyleSheet() atau styles kustom.
    """
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import Paragraph, Spacer, Table, TableStyle

    name_style = ParagraphStyle(
        "CompanyName", parent=styles["Normal"], fontSize=14,
        fontName="Helvetica-Bold", textColor=colors.HexColor("#0d6efd"),
    )
    info_style = ParagraphStyle(
        "CompanyInfo", parent=styles["Normal"], fontSize=8.5,
        textColor=colors.HexColor("#333333"), leading=11,
    )

    # Susun baris info perusahaan
    info_lines = [f"<b>{company.get('company_name', '')}</b>"]
    if company.get("address"):
        info_lines.append(company["address"])
    contact = []
    if company.get("phone"):
        contact.append(f"Telp: {company['phone']}")
    if company.get("email"):
        contact.append(company["email"])
    if contact:
        info_lines.append(" &nbsp;|&nbsp; ".join(contact))
    if company.get("npwp"):
        info_lines.append(f"NPWP: {company['npwp']}")

    info_para = Paragraph("<br/>".join(info_lines), info_style)
    logo = _try_load_logo(company.get("logo_path", ""))

    if logo is not None:
        # Dua kolom: logo kiri, info kanan
        head_table = Table([[logo, info_para]], colWidths=[3 * cm, 14 * cm])
        head_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ]))
        elements.append(head_table)
    else:
        elements.append(Paragraph(company.get("company_name", ""), name_style))
        elements.append(info_para)

    # Garis pemisah
    elements.append(Spacer(1, 0.15 * cm))
    line = Table([[""]], colWidths=[17.5 * cm], rowHeights=[2])
    line.setStyle(TableStyle([("LINEBELOW", (0, 0), (-1, -1), 1, colors.HexColor("#0d6efd"))]))
    elements.append(line)
    elements.append(Spacer(1, 0.4 * cm))


def build_bank_footer(elements: list, company: dict, styles) -> None:
    """Tambahkan footer rekening bank (untuk invoice) bila datanya ada."""
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import Paragraph, Spacer

    if not company.get("bank_name") and not company.get("account_number"):
        return
    foot_style = ParagraphStyle(
        "BankFooter", parent=styles["Normal"], fontSize=8.5,
        textColor=colors.HexColor("#333333"), leading=11,
    )
    parts = ["<b>Pembayaran ditransfer ke:</b>"]
    if company.get("bank_name"):
        parts.append(f"Bank: {company['bank_name']}")
    if company.get("account_number"):
        parts.append(f"No. Rekening: {company['account_number']}")
    if company.get("account_holder"):
        parts.append(f"Atas Nama: {company['account_holder']}")
    elements.append(Spacer(1, 0.4 * cm))
    elements.append(Paragraph("<br/>".join(parts), foot_style))


def add_excel_letterhead(ws, company: dict, num_cols: int = 6) -> int:
    """Tambahkan kop surat ke worksheet openpyxl di bagian atas.

    Menulis logo (bila ada) + nama & info perusahaan, lalu mengembalikan nomor
    baris pertama yang BEBAS dipakai untuk konten setelah kop (mis. judul tabel).

    Parameters
    ----------
    ws       : worksheet openpyxl.
    company  : hasil get_company_settings().
    num_cols : jumlah kolom tabel (untuk merge teks kop).

    Returns
    -------
    int : nomor baris berikutnya yang bebas (>= 1).
    """
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    last_col = get_column_letter(num_cols)

    # Coba pasang logo: dukung file lokal (logo bundel) ATAU URL (logo upload).
    logo_added = False
    logo_path = company.get("logo_path", "")
    try:
        import os
        from openpyxl.drawing.image import Image as XLImage

        img_source = None
        if logo_path and str(logo_path).startswith(("http://", "https://")):
            # Logo dari URL (mis. hasil upload ke Supabase Storage): unduh dulu.
            import urllib.request
            with urllib.request.urlopen(str(logo_path), timeout=5) as resp:
                img_source = BytesIO(resp.read())
        elif logo_path and os.path.exists(str(logo_path)):
            img_source = str(logo_path)

        if img_source is not None:
            img = XLImage(img_source)
            # Skala logo ~70px tinggi menjaga rasio
            target_h = 70
            if img.height:
                ratio = img.width / img.height
                img.height = target_h
                img.width = int(target_h * ratio)
            ws.add_image(img, "A1")
            logo_added = True
    except Exception:
        logo_added = False

    # Teks kop di sebelah kanan logo (mulai kolom B bila ada logo, else A)
    text_col_start = 2 if logo_added else 1
    text_col_letter = get_column_letter(text_col_start)

    name = company.get("company_name", "")
    addr = company.get("address", "")
    contact_bits = []
    if company.get("phone"):
        contact_bits.append(f"Telp: {company['phone']}")
    if company.get("email"):
        contact_bits.append(company["email"])
    contact = "  |  ".join(contact_bits)
    npwp = f"NPWP: {company['npwp']}" if company.get("npwp") else ""

    # Baris 1: nama perusahaan (bold, biru)
    ws.merge_cells(f"{text_col_letter}1:{last_col}1")
    c = ws[f"{text_col_letter}1"]
    c.value = name
    c.font = Font(bold=True, size=14, color="0d6efd")
    c.alignment = Alignment(horizontal="left", vertical="center")

    # Baris 2: alamat
    ws.merge_cells(f"{text_col_letter}2:{last_col}2")
    ws[f"{text_col_letter}2"].value = addr
    ws[f"{text_col_letter}2"].font = Font(size=9)
    ws[f"{text_col_letter}2"].alignment = Alignment(horizontal="left")

    # Baris 3: kontak + NPWP
    line3 = contact + (("  |  " + npwp) if (contact and npwp) else npwp)
    ws.merge_cells(f"{text_col_letter}3:{last_col}3")
    ws[f"{text_col_letter}3"].value = line3
    ws[f"{text_col_letter}3"].font = Font(size=9)

    # Beri tinggi baris agar logo muat
    for r in (1, 2, 3):
        ws.row_dimensions[r].height = 22

    # Baris 4 dibiarkan kosong sebagai pemisah; konten mulai baris 5.
    return 5


def upload_logo(file_bytes: bytes, filename: str, content_type: str = "image/png") -> str:
    """Upload logo ke Supabase Storage (bucket 'opname-photos', folder company/).

    Memakai bucket yang sama dengan foto opname (sudah ada & berfungsi).
    Mengembalikan public URL logo. Melempar exception bila gagal (agar pemanggil
    bisa menampilkan error).
    """
    import uuid

    supabase = get_supabase()
    ext = (filename.rsplit(".", 1)[-1] if "." in filename else "png").lower()
    path = f"company/logo_{uuid.uuid4().hex}.{ext}"
    supabase.storage.from_("opname-photos").upload(
        path=path,
        file=file_bytes,
        file_options={"content-type": content_type},
    )
    return supabase.storage.from_("opname-photos").get_public_url(path)
